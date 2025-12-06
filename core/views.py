from django.shortcuts import render, redirect, get_object_or_404
from django.db import connection, transaction, models
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, FileResponse
from .models import *
from .validators import *
import pandas as pd
from datetime import datetime, date
from django.conf import settings
from io import BytesIO
import os
import openpyxl
import openpyxl.utils
import json
import numpy as np
import pandas as pd
import re
import subprocess
import tempfile



def home(request):
    return render(request, 'core/home.html')


def entities_menu(request):
    entity = request.GET.get('entity', '')
    operation = request.GET.get('operation', '')

    if entity and operation:
        if operation == 'view':
            return redirect(f'{entity}_list')
        elif operation == 'add':
            return redirect(f'{entity}_add')
        elif operation == 'update':
            return redirect(f'{entity}_update')
        elif operation == 'delete':
            return redirect(f'{entity}_delete')

    return render(request, 'core/entities_menu.html')


# ============= VISITOR CRUD =============
def visitor_list(request):
    visitors = Visitor.objects.all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        filter_dict = {filter_column + '__icontains': filter_value}
        visitors = visitors.filter(**filter_dict)

    if request.GET.get('save_result'):
        return save_query_result(list(visitors.values()), 'visitors_filtered')

    columns = ['visitor_id', 'full_name', 'birth_date', 'citizenship', 'ticket_type', 'visit_date', 'review']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Посетители',
        'items': visitors,
        'columns': columns,
        'entity_type': 'visitor',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def visitor_add(request):
    if request.method == 'POST':
        try:
            # Получаем максимальный ID
            max_id = Visitor.objects.aggregate(models.Max('visitor_id'))['visitor_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(Visitor, 'visitor_id', new_id)

            # Валидация ФИО
            full_name = validate_fio(request.POST.get('full_name', ''))

            # Валидация дат
            birth_date_str = request.POST.get('birth_date', '')
            visit_date_str = request.POST.get('visit_date', '')

            # Преобразуем из формата ДД.ММ.ГГГГ если нужно
            if '.' in birth_date_str:
                birth_date = validate_date_format(birth_date_str)
            else:
                birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()

            if '.' in visit_date_str:
                visit_date = validate_date_format(visit_date_str)
            else:
                visit_date = datetime.strptime(visit_date_str, '%Y-%m-%d').date()

            # Логические проверки дат
            if birth_date > date.today():
                messages.error(request, 'Дата рождения не может быть в будущем!')
                return redirect('visitor_add')

            if visit_date < birth_date:
                messages.error(request, 'Дата посещения не может быть раньше даты рождения!')
                return redirect('visitor_add')

            if visit_date > date.today():
                messages.error(request, 'Дата посещения не может быть в будущем!')
                return redirect('visitor_add')

            # Валидация гражданства
            citizenship = request.POST.get('citizenship', '')
            if citizenship not in COUNTRIES:
                messages.error(request, 'Выберите гражданство из списка!')
                return redirect('visitor_add')

            # Валидация типа билета
            ticket_type = request.POST.get('ticket_type', '')
            if ticket_type not in TICKET_TYPES:
                messages.error(request, 'Выберите тип билета из списка!')
                return redirect('visitor_add')

            # Валидация отзыва
            review = request.POST.get('review', '').strip()
            if review:
                if len(review) > 100:
                    messages.error(request, 'Отзыв не может быть длиннее 100 символов!')
                    return redirect('visitor_add')
                # Проверка на бессмысленные символы
                if re.match(r'^[!@#$%^&*()_+=\-\s]+$', review):
                    messages.error(request, 'Отзыв не может состоять только из специальных символов!')
                    return redirect('visitor_add')

            visitor = Visitor(
                visitor_id=new_id,
                full_name=full_name,
                birth_date=birth_date,
                citizenship=citizenship,
                ticket_type=ticket_type,
                visit_date=visit_date,
                review=review if review else ''
            )

            visitor.save()
            messages.success(request, 'Посетитель успешно добавлен!')
            return redirect('visitor_list')

        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('visitor_add')
        except Exception as e:
            messages.error(request, f'Ошибка при добавлении: {str(e)}')
            return redirect('visitor_add')

    # Подготовка данных для формы
    fields_data = {
        'countries': COUNTRIES,
        'ticket_types': TICKET_TYPES,
    }

    return render(request, 'core/visitor_form_add.html', {
        'entity_name': 'Посетитель',
        'operation': 'Добавить',
        'entity_type': 'visitor',
        'fields_data': fields_data
    })


def visitor_update(request):
    visitors = Visitor.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        visitor = get_object_or_404(Visitor, visitor_id=selected_id)

        if request.method == 'POST':
            try:
                # Валидация ФИО
                full_name = validate_fio(request.POST.get('full_name', ''))

                # Валидация дат
                birth_date_str = request.POST.get('birth_date', '')
                visit_date_str = request.POST.get('visit_date', '')

                # Преобразуем из формата ДД.ММ.ГГГГ если нужно
                if '.' in birth_date_str:
                    birth_date = validate_date_format(birth_date_str)
                else:
                    birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()

                if '.' in visit_date_str:
                    visit_date = validate_date_format(visit_date_str)
                else:
                    visit_date = datetime.strptime(visit_date_str, '%Y-%m-%d').date()

                # Логические проверки дат
                if birth_date > date.today():
                    messages.error(request, 'Дата рождения не может быть в будущем!')
                    return redirect(f'visitor_update?id={selected_id}')

                if visit_date < birth_date:
                    messages.error(request, 'Дата посещения не может быть раньше даты рождения!')
                    return redirect(f'visitor_update?id={selected_id}')

                if visit_date > date.today():
                    messages.error(request, 'Дата посещения не может быть в будущем!')
                    return redirect(f'visitor_update?id={selected_id}')

                # Валидация гражданства
                citizenship = request.POST.get('citizenship', '')
                if citizenship not in COUNTRIES:
                    messages.error(request, 'Выберите гражданство из списка!')
                    return redirect(f'visitor_update?id={selected_id}')

                # Валидация типа билета
                ticket_type = request.POST.get('ticket_type', '')
                if ticket_type not in TICKET_TYPES:
                    messages.error(request, 'Выберите тип билета из списка!')
                    return redirect(f'visitor_update?id={selected_id}')

                # Валидация отзыва
                review = request.POST.get('review', '').strip()
                if review:
                    if len(review) > 100:
                        messages.error(request, 'Отзыв не может быть длиннее 100 символов!')
                        return redirect(f'visitor_update?id={selected_id}')
                    if re.match(r'^[!@#$%^&*()_+=\-\s]+$', review):
                        messages.error(request, 'Отзыв не может состоять только из специальных символов!')
                        return redirect(f'visitor_update?id={selected_id}')

                # Обновляем данные
                visitor.full_name = full_name
                visitor.birth_date = birth_date
                visitor.visit_date = visit_date
                visitor.citizenship = citizenship
                visitor.ticket_type = ticket_type
                visitor.review = review if review else ''

                visitor.save()
                messages.success(request, 'Посетитель успешно обновлен!')
                return redirect('visitor_list')

            except ValidationError as e:
                messages.error(request, str(e))
                return redirect(f'visitor_update?id={selected_id}')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')
                return redirect(f'visitor_update?id={selected_id}')

        # Подготовка данных для формы
        fields_data = {
            'countries': COUNTRIES,
            'ticket_types': TICKET_TYPES,
            'visitor': visitor
        }

        return render(request, 'core/visitor_form_update.html', {
            'entity_name': 'Посетитель',
            'operation': 'Обновить',
            'entity_type': 'visitor',
            'fields_data': fields_data
        })

    columns = ['visitor_id', 'full_name', 'birth_date', 'citizenship']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Посетитель',
        'items': visitors,
        'columns': columns,
        'entity_type': 'visitor',
        'operation': 'обновления'
    })

    columns = ['visitor_id', 'full_name', 'birth_date', 'citizenship']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Посетитель',
        'items': visitors,
        'columns': columns,
        'entity_type': 'visitor',
        'operation': 'обновления'
    })


def visitor_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                visitor = Visitor.objects.get(visitor_id=delete_id)
                visitor.delete()
                messages.success(request, f'Посетитель "{visitor.full_name}" успешно удален!')
            except Visitor.DoesNotExist:
                messages.error(request, 'Посетитель не найден!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('visitor_delete')

    visitors = Visitor.objects.all()
    columns = ['visitor_id', 'full_name', 'birth_date', 'citizenship']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Посетитель',
        'items': visitors,
        'columns': columns,
        'entity_type': 'visitor',
        'operation': 'удаления'
    })


# ============= EMPLOYEE CRUD =============
def employee_list(request):
    employees = Employee.objects.all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        filter_dict = {filter_column + '__icontains': filter_value}
        employees = employees.filter(**filter_dict)

    if request.GET.get('save_result'):
        return save_query_result(list(employees.values()), 'employees_filtered')

    columns = ['employee_id', 'full_name', 'position', 'hire_date', 'department', 'phone', 'qualification']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Сотрудники',
        'items': employees,
        'columns': columns,
        'entity_type': 'employee',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def employee_add(request):
    # Инициализируем form_data и fields_data ДО условия POST
    form_data = {}

    if request.method == 'POST':
        # Сохраняем все введенные данные
        form_data = {
            'full_name': request.POST.get('full_name', ''),
            'department': request.POST.get('department', ''),
            'position': request.POST.get('position', ''),
            'hire_date': request.POST.get('hire_date', ''),
            'phone_code': request.POST.get('phone_code', '+7'),
            'phone_number': request.POST.get('phone_number', ''),
            'qualification': request.POST.get('qualification', ''),
        }

        try:
            # Получаем максимальный ID
            max_id = Employee.objects.aggregate(models.Max('employee_id'))['employee_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(Employee, 'employee_id', new_id)

            # Валидация ФИО
            full_name = validate_fio(form_data['full_name'])

            # Валидация отдела
            department = form_data['department']
            if department not in DEPARTMENTS:
                error = FieldValidationError('Выберите отдел из списка!')
                error.field = 'department'
                raise error

            # Валидация должности
            position = form_data['position']
            if position not in POSITIONS_BY_DEPARTMENT.get(department, []):
                error = FieldValidationError(f'Выберите корректную должность для отдела {department}!')
                error.field = 'position'
                raise error

            # Валидация даты найма
            hire_date_str = form_data['hire_date']
            if '.' in hire_date_str:
                hire_date = validate_date_format(hire_date_str)
            else:
                hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()

            if hire_date > date.today():
                error = FieldValidationError('Дата найма не может быть в будущем!')
                error.field = 'hire_date'
                raise error

            if hire_date < date(1990, 1, 1):
                error = FieldValidationError('Дата найма не может быть раньше 1990 года!')
                error.field = 'hire_date'
                raise error

            # Валидация телефона
            country_code = form_data['phone_code']
            phone_number = form_data['phone_number']
            phone = validate_phone(phone_number, country_code)

            # Проверка уникальности телефона
            check_unique_phone(Employee, phone)

            # Валидация квалификации
            qualification = validate_text_field(
                form_data['qualification'],
                max_length=70,
                min_length=2,
                allow_digits=False,
                allow_special=False
            )

            employee = Employee(
                employee_id=new_id,
                full_name=full_name,
                position=position,
                hire_date=hire_date,
                department=department,
                phone=phone,
                qualification=qualification
            )

            employee.save()
            messages.success(request, 'Сотрудник успешно добавлен!')
            return redirect('employee_list')

        except ValidationError as e:
            messages.error(request, str(e))
            fields_data = {
                'departments': DEPARTMENTS,
                'positions_by_department': POSITIONS_BY_DEPARTMENT,
                'form_data': form_data,
                'error_field': e.field if hasattr(e, 'field') else None
            }
            return render(request, 'core/employee_form_add.html', {
                'entity_name': 'Сотрудник',
                'operation': 'Добавить',
                'entity_type': 'employee',
                'fields_data': fields_data
            })

    # При GET запросе или если не было исключения
    fields_data = {
        'departments': DEPARTMENTS,
        'positions_by_department': POSITIONS_BY_DEPARTMENT,
        'form_data': form_data,  # Будет пустым при GET
        'error_field': None
    }

    return render(request, 'core/employee_form_add.html', {
        'entity_name': 'Сотрудник',
        'operation': 'Добавить',
        'entity_type': 'employee',
        'fields_data': fields_data
    })


def employee_update(request):
    employees = Employee.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        employee = get_object_or_404(Employee, employee_id=selected_id)

        if request.method == 'POST':
            try:
                # Валидация ФИО
                full_name = validate_fio(request.POST.get('full_name', ''))

                # Валидация отдела
                department = request.POST.get('department', '')
                if department not in DEPARTMENTS:
                    messages.error(request, 'Выберите отдел из списка!')
                    return redirect(f'employee_update?id={selected_id}')

                # Валидация должности
                position = request.POST.get('position', '')
                if position not in POSITIONS_BY_DEPARTMENT.get(department, []):
                    messages.error(request, f'Выберите корректную должность для отдела {department}!')
                    return redirect(f'employee_update?id={selected_id}')

                # Валидация даты найма
                hire_date_str = request.POST.get('hire_date', '')
                if '.' in hire_date_str:
                    hire_date = validate_date_format(hire_date_str)
                else:
                    hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()

                if hire_date > date.today():
                    messages.error(request, 'Дата найма не может быть в будущем!')
                    return redirect(f'employee_update?id={selected_id}')

                if hire_date < date(1990, 1, 1):
                    messages.error(request, 'Дата найма не может быть раньше 1990 года!')
                    return redirect(f'employee_update?id={selected_id}')

                # Валидация телефона
                country_code = request.POST.get('phone_code', '')
                phone_number = request.POST.get('phone_number', '')
                phone = validate_phone(phone_number, country_code)

                # Проверка уникальности телефона (исключая текущего сотрудника)
                check_unique_phone(Employee, phone, exclude_id=selected_id)

                # Валидация квалификации
                qualification = validate_text_field(
                    request.POST.get('qualification', ''),
                    max_length=70,
                    min_length=2,
                    allow_digits=False,
                    allow_special=False
                )

                # Обновляем данные
                employee.full_name = full_name
                employee.position = position
                employee.hire_date = hire_date
                employee.department = department
                employee.phone = phone
                employee.qualification = qualification

                employee.save()
                messages.success(request, 'Сотрудник успешно обновлен!')
                return redirect('employee_list')

            except ValidationError as e:
                messages.error(request, str(e))
                return redirect(f'employee_update?id={selected_id}')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')
                return redirect(f'employee_update?id={selected_id}')

        # Разделяем телефон на код и номер для отображения
        phone_code = ''
        phone_number = employee.phone
        if employee.phone.startswith('+375'):
            phone_code = '+375'
            phone_number = employee.phone[4:]
        elif employee.phone.startswith('+7'):
            phone_code = '+7'
            phone_number = employee.phone[2:]

        fields_data = {
            'departments': DEPARTMENTS,
            'positions_by_department': POSITIONS_BY_DEPARTMENT,
            'employee': employee,
            'phone_code': phone_code,
            'phone_number': phone_number
        }

        return render(request, 'core/employee_form_update.html', {
            'entity_name': 'Сотрудник',
            'operation': 'Обновить',
            'entity_type': 'employee',
            'fields_data': fields_data
        })

    columns = ['employee_id', 'full_name', 'position', 'department']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Сотрудник',
        'items': employees,
        'columns': columns,
        'entity_type': 'employee',
        'operation': 'обновления'
    })


def employee_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                employee = Employee.objects.get(employee_id=delete_id)
                employee.delete()
                messages.success(request, f'Сотрудник "{employee.full_name}" успешно удален!')
            except Employee.DoesNotExist:
                messages.error(request, 'Сотрудник не найден!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('employee_delete')

    employees = Employee.objects.all()
    columns = ['employee_id', 'full_name', 'position', 'department']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Сотрудник',
        'items': employees,
        'columns': columns,
        'entity_type': 'employee',
        'operation': 'удаления'
    })


# ============= EXHIBITION CRUD =============
def exhibition_list(request):
    exhibitions = Exhibition.objects.all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        filter_dict = {filter_column + '__icontains': filter_value}
        exhibitions = exhibitions.filter(**filter_dict)

    if request.GET.get('save_result'):
        return save_query_result(list(exhibitions.values()), 'exhibitions_filtered')

    columns = ['exhibition_id', 'title', 'theme', 'start_date', 'end_date', 'location', 'type']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Выставки',
        'items': exhibitions,
        'columns': columns,
        'entity_type': 'exhibition',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def exhibition_add(request):
    form_data = {}

    if request.method == 'POST':
        form_data = {
            'title': request.POST.get('title', ''),
            'theme': request.POST.get('theme', ''),
            'start_date': request.POST.get('start_date', ''),
            'end_date': request.POST.get('end_date', ''),
            'location': request.POST.get('location', ''),
            'type': request.POST.get('type', ''),
        }

        try:
            # Получаем максимальный ID
            max_id = Exhibition.objects.aggregate(models.Max('exhibition_id'))['exhibition_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(Exhibition, 'exhibition_id', new_id)

            # Валидация названия
            title = validate_text_field(
                form_data['title'],
                max_length=50,
                min_length=2,
                allow_digits=False,
                allow_special=False
            )

            # Валидация темы
            theme = validate_text_field(
                form_data['theme'],
                max_length=50,
                min_length=2,
                allow_digits=False,
                allow_special=False
            )

            # Валидация дат
            if '.' in form_data['start_date']:
                start_date = validate_date_format(form_data['start_date'])
            else:
                start_date = datetime.strptime(form_data['start_date'], '%Y-%m-%d').date()

            if '.' in form_data['end_date']:
                end_date = validate_date_format(form_data['end_date'])
            else:
                end_date = datetime.strptime(form_data['end_date'], '%Y-%m-%d').date()

            # Логические проверки дат
            if end_date < start_date:
                error = FieldValidationError('Дата окончания не может быть раньше даты начала!')
                error.field = 'end_date'
                raise error

            # Валидация места
            if form_data['location'] not in EXHIBITION_LOCATIONS:
                error = FieldValidationError('Выберите место из списка!')
                error.field = 'location'
                raise error

            # Валидация типа
            if form_data['type'] not in EXHIBITION_TYPES:
                error = FieldValidationError('Выберите тип выставки!')
                error.field = 'type'
                raise error

            exhibition = Exhibition(
                exhibition_id=new_id,
                title=title,
                theme=theme,
                start_date=start_date,
                end_date=end_date,
                location=form_data['location'],
                type=form_data['type']
            )

            exhibition.save()
            messages.success(request, 'Выставка успешно добавлена!')
            return redirect('exhibition_list')

        except ValidationError as e:
            messages.error(request, str(e))
            fields_data = {
                'locations': EXHIBITION_LOCATIONS,
                'types': EXHIBITION_TYPES,
                'form_data': form_data,
                'error_field': e.field if hasattr(e, 'field') else None
            }
            return render(request, 'core/exhibition_form_add.html', {
                'entity_name': 'Выставка',
                'operation': 'Добавить',
                'entity_type': 'exhibition',
                'fields_data': fields_data
            })

    fields_data = {
        'locations': EXHIBITION_LOCATIONS,
        'types': EXHIBITION_TYPES,
        'form_data': form_data
    }

    return render(request, 'core/exhibition_form_add.html', {
        'entity_name': 'Выставка',
        'operation': 'Добавить',
        'entity_type': 'exhibition',
        'fields_data': fields_data
    })


def exhibition_update(request):
    exhibitions = Exhibition.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        exhibition = get_object_or_404(Exhibition, exhibition_id=selected_id)

        if request.method == 'POST':
            try:
                exhibition.title = request.POST['title']
                exhibition.theme = request.POST['theme']

                start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()
                end_date = datetime.strptime(request.POST['end_date'], '%Y-%m-%d').date()

                # Валидация
                if end_date < start_date:
                    messages.error(request, 'Дата начала не может быть больше даты окончания!')
                    return redirect(f'exhibition_update?id={selected_id}')

                exhibition.start_date = start_date
                exhibition.end_date = end_date
                exhibition.location = request.POST['location']
                exhibition.type = request.POST['type']

                exhibition.save()
                messages.success(request, 'Выставка успешно обновлена!')
                return redirect('exhibition_list')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')

        fields = [
            {'name': 'title', 'label': 'Название', 'type': 'text', 'required': True, 'value': exhibition.title},
            {'name': 'theme', 'label': 'Тема', 'type': 'text', 'required': True, 'value': exhibition.theme},
            {'name': 'start_date', 'label': 'Дата начала', 'type': 'date', 'required': True,
             'value': exhibition.start_date},
            {'name': 'end_date', 'label': 'Дата окончания', 'type': 'date', 'required': True,
             'value': exhibition.end_date},
            {'name': 'location', 'label': 'Место проведения', 'type': 'text', 'required': True,
             'value': exhibition.location},
            {'name': 'type', 'label': 'Тип', 'type': 'text', 'required': True, 'value': exhibition.type},
        ]

        return render(request, 'core/entity_form_update.html', {
            'entity_name': 'Выставка',
            'operation': 'Обновить',
            'fields': fields,
            'entity_type': 'exhibition'
        })

    columns = ['exhibition_id', 'title', 'theme', 'location']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Выставка',
        'items': exhibitions,
        'columns': columns,
        'entity_type': 'exhibition',
        'operation': 'обновления'
    })


def exhibition_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                exhibition = Exhibition.objects.get(exhibition_id=delete_id)
                exhibition.delete()
                messages.success(request, f'Выставка "{exhibition.title}" успешно удалена!')
            except Exhibition.DoesNotExist:
                messages.error(request, 'Выставка не найдена!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('exhibition_delete')

    exhibitions = Exhibition.objects.all()
    columns = ['exhibition_id', 'title', 'theme', 'location']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Выставка',
        'items': exhibitions,
        'columns': columns,
        'entity_type': 'exhibition',
        'operation': 'удаления'
    })


# ============= EXCURSION CRUD =============
def excursion_list(request):
    # Аналогично для экскурсий
    excursions = Excursion.objects.select_related('employee').all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        filter_dict = {filter_column + '__icontains': filter_value}
        excursions = excursions.filter(**filter_dict)

    if request.GET.get('save_result'):
        data = []
        for excursion in excursions:
            data.append({
                'excursion_id': excursion.excursion_id,
                'title': excursion.title,
                'date': excursion.date,
                'language': excursion.language,
                'ticket_num': excursion.ticket_num,
                'price': excursion.price,
                'duration': excursion.duration,
                'employee': excursion.employee.full_name if excursion.employee else None
            })
        return save_query_result(data, 'excursions_filtered')

    # Подготавливаем данные для отображения
    excursions_display = []
    for excursion in excursions:
        excursions_display.append({
            'excursion_id': excursion.excursion_id,
            'title': excursion.title,
            'date': excursion.date,
            'language': excursion.language,
            'ticket_num': excursion.ticket_num,
            'price': excursion.price,
            'duration': excursion.duration,
            'employee_name': excursion.employee.full_name if excursion.employee else None
        })

    columns = ['excursion_id', 'title', 'date', 'language', 'ticket_num', 'price', 'duration', 'employee_id']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Экскурсии',
        'items': excursions_display,
        'columns': columns,
        'entity_type': 'excursion',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def excursion_add(request):
    form_data = {}

    if request.method == 'POST':
        form_data = {
            'title': request.POST.get('title', ''),
            'date': request.POST.get('date', ''),
            'language': request.POST.get('language', ''),
            'ticket_num': request.POST.get('ticket_num', ''),
            'price': request.POST.get('price', ''),
            'duration': request.POST.get('duration', ''),
            'employee_id': request.POST.get('employee_id', ''),
        }

        try:
            # Получаем максимальный ID
            max_id = Excursion.objects.aggregate(models.Max('excursion_id'))['excursion_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(Excursion, 'excursion_id', new_id)

            # Валидация названия
            title = validate_text_field(
                form_data['title'],
                max_length=50,
                min_length=2,
                allow_digits=False,
                allow_special=False
            )

            # Валидация даты
            if '.' in form_data['date']:
                excursion_date = validate_date_format(form_data['date'])
            else:
                excursion_date = datetime.strptime(form_data['date'], '%Y-%m-%d').date()

            # Валидация языка
            if form_data['language'] not in EXCURSION_LANGUAGES:
                error = FieldValidationError('Выберите язык из списка!')
                error.field = 'language'
                raise error

            # Валидация количества билетов
            try:
                ticket_num = int(form_data['ticket_num'])
                if ticket_num < 10 or ticket_num > 100:
                    error = FieldValidationError('Количество билетов должно быть от 10 до 100!')
                    error.field = 'ticket_num'
                    raise error
            except ValueError:
                error = FieldValidationError('Введите корректное число билетов!')
                error.field = 'ticket_num'
                raise error

            # Валидация цены
            try:
                price = float(form_data['price'])
                if price < 20 or price > 300:
                    error = FieldValidationError('Цена должна быть от 20 до 300!')
                    error.field = 'price'
                    raise error
            except ValueError:
                error = FieldValidationError('Введите корректную цену!')
                error.field = 'price'
                raise error

            # Валидация продолжительности
            try:
                duration = int(form_data['duration'])
                if duration < 30 or duration > 180:
                    error = FieldValidationError('Продолжительность должна быть от 30 до 180 минут!')
                    error.field = 'duration'
                    raise error
            except ValueError:
                error = FieldValidationError('Введите корректную продолжительность!')
                error.field = 'duration'
                raise error

            excursion = Excursion(
                excursion_id=new_id,
                title=title,
                date=excursion_date,
                language=form_data['language'],
                ticket_num=ticket_num,
                price=price,
                duration=duration
            )

            # Обработка employee_id
            if form_data['employee_id'].strip():
                try:
                    employee = Employee.objects.get(employee_id=int(form_data['employee_id']))
                    excursion.employee = employee
                except (ValueError, Employee.DoesNotExist):
                    error = FieldValidationError(f'Сотрудник с ID {form_data["employee_id"]} не найден!')
                    error.field = 'employee_id'
                    raise error

            excursion.save()
            messages.success(request, 'Экскурсия успешно добавлена!')
            return redirect('excursion_list')

        except ValidationError as e:
            messages.error(request, str(e))
            fields_data = {
                'languages': EXCURSION_LANGUAGES,
                'form_data': form_data,
                'error_field': e.field if hasattr(e, 'field') else None
            }
            return render(request, 'core/excursion_form_add.html', {
                'entity_name': 'Экскурсия',
                'operation': 'Добавить',
                'entity_type': 'excursion',
                'fields_data': fields_data
            })

    fields_data = {
        'languages': EXCURSION_LANGUAGES,
        'form_data': form_data
    }

    return render(request, 'core/excursion_form_add.html', {
        'entity_name': 'Экскурсия',
        'operation': 'Добавить',
        'entity_type': 'excursion',
        'fields_data': fields_data
    })


def excursion_update(request):
    excursions = Excursion.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        excursion = get_object_or_404(Excursion, excursion_id=selected_id)

        if request.method == 'POST':
            try:
                excursion.title = request.POST['title']

                # Преобразуем строку в дату
                excursion_date = datetime.strptime(request.POST['date'], '%Y-%m-%d').date()
                excursion.date = excursion_date

                excursion.language = request.POST['language']
                excursion.ticket_num = request.POST['ticket_num']
                excursion.price = request.POST['price']
                excursion.duration = request.POST['duration']

                employee_id_value = request.POST.get('employee_id', '').strip()
                if employee_id_value:
                    try:
                        employee = Employee.objects.get(employee_id=int(employee_id_value))
                        excursion.employee = employee
                    except (ValueError, Employee.DoesNotExist):
                        messages.error(request, f'Сотрудник с ID {employee_id_value} не найден!')
                        return redirect(f'excursion_update?id={selected_id}')
                else:
                    excursion.employee = None

                # Валидация
                if int(excursion.ticket_num) < 0:
                    messages.error(request, 'Количество билетов не может быть отрицательным!')
                    return redirect(f'excursion_update?id={selected_id}')

                if float(excursion.price) < 0:
                    messages.error(request, 'Цена не может быть отрицательной!')
                    return redirect(f'excursion_update?id={selected_id}')

                if int(excursion.duration) <= 0:
                    messages.error(request, 'Продолжительность должна быть положительной!')
                    return redirect(f'excursion_update?id={selected_id}')

                excursion.save()
                messages.success(request, 'Экскурсия успешно обновлена!')
                return redirect('excursion_list')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')

        fields = [
            {'name': 'title', 'label': 'Название', 'type': 'text', 'required': True, 'value': excursion.title},
            {'name': 'date', 'label': 'Дата', 'type': 'date', 'required': True, 'value': excursion.date},
            {'name': 'language', 'label': 'Язык', 'type': 'text', 'required': True, 'value': excursion.language},
            {'name': 'ticket_num', 'label': 'Количество билетов', 'type': 'number', 'required': True,
             'value': excursion.ticket_num},
            {'name': 'price', 'label': 'Цена', 'type': 'number', 'required': True, 'step': '0.01',
             'value': excursion.price},
            {'name': 'duration', 'label': 'Продолжительность (минуты)', 'type': 'number', 'required': True,
             'value': excursion.duration},
            {'name': 'employee_id', 'label': 'ID сотрудника (необязательно)', 'type': 'number', 'required': False,
             'value': excursion.employee_id if excursion.employee else ''},
        ]

        return render(request, 'core/entity_form_update.html', {
            'entity_name': 'Экскурсия',
            'operation': 'Обновить',
            'fields': fields,
            'entity_type': 'excursion'
        })

    columns = ['excursion_id', 'title', 'date', 'language']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Экскурсия',
        'items': excursions,
        'columns': columns,
        'entity_type': 'excursion',
        'operation': 'обновления'
    })


def excursion_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                excursion = Excursion.objects.get(excursion_id=delete_id)
                excursion.delete()
                messages.success(request, f'Экскурсия "{excursion.title}" успешно удалена!')
            except Excursion.DoesNotExist:
                messages.error(request, 'Экскурсия не найдена!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('excursion_delete')

    excursions = Excursion.objects.all()
    columns = ['excursion_id', 'title', 'date', 'language']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Экскурсия',
        'items': excursions,
        'columns': columns,
        'entity_type': 'excursion',
        'operation': 'удаления'
    })


# ============= EXHIBIT CRUD =============
def exhibit_list(request):
    # Используем правильное имя поля - mission_id
    exhibits = Exhibit.objects.select_related('mission_id').all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        filter_dict = {filter_column + '__icontains': filter_value}
        exhibits = exhibits.filter(**filter_dict)

    if request.GET.get('save_result'):
        data = []
        for exhibit in exhibits:
            data.append({
                'exhibit_id': exhibit.exhibit_id,
                'title': exhibit.title,
                'description': exhibit.description,
                'creation_date': exhibit.creation_date,
                'country': exhibit.country,
                'state': exhibit.state,
                'type': exhibit.type,
                'mission': exhibit.mission_id.title if exhibit.mission_id else None
            })
        return save_query_result(data, 'exhibits_filtered')

    # Подготавливаем данные для отображения
    exhibits_display = []
    for exhibit in exhibits:
        exhibits_display.append({
            'exhibit_id': exhibit.exhibit_id,
            'title': exhibit.title,
            'description': exhibit.description,
            'creation_date': exhibit.creation_date,
            'country': exhibit.country,
            'state': exhibit.state,
            'type': exhibit.type,
            'mission_title': exhibit.mission_id.title if exhibit.mission_id else None
        })

    columns = ['exhibit_id', 'title', 'description', 'creation_date', 'country', 'state', 'type', 'mission_id']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Экспонаты',
        'items': exhibits_display,
        'columns': columns,
        'entity_type': 'exhibit',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def exhibit_add(request):
    form_data = {}

    if request.method == 'POST':
        form_data = {
            'title': request.POST.get('title', ''),
            'description': request.POST.get('description', ''),
            'creation_date': request.POST.get('creation_date', ''),
            'country': request.POST.get('country', ''),
            'state': request.POST.get('state', ''),
            'type': request.POST.get('type', ''),
            'mission_id': request.POST.get('mission_id', ''),
        }

        try:
            # Получаем максимальный ID
            max_id = Exhibit.objects.aggregate(models.Max('exhibit_id'))['exhibit_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(Exhibit, 'exhibit_id', new_id)

            # Валидация названия (разрешаем цифры, скобки, кавычки)
            title = form_data['title'].strip()
            if len(title) < 2:
                error = FieldValidationError('Название должно содержать минимум 2 символа')
                error.field = 'title'
                raise error
            if len(title) > 70:
                error = FieldValidationError('Название не может быть длиннее 70 символов')
                error.field = 'title'
                raise error
            # Проверяем на недопустимые символы (разрешаем буквы, цифры, пробелы, скобки, кавычки, дефис)
            if not re.match(r'^[а-яА-ЯёЁa-zA-Z0-9\s\(\)\"\-]+$', title):
                error = FieldValidationError('Название содержит недопустимые символы')
                error.field = 'title'
                raise error

            # Валидация описания
            description = form_data['description'].strip()
            if len(description) < 5:
                error = FieldValidationError('Описание должно содержать минимум 5 символов')
                error.field = 'description'
                raise error
            if len(description) > 100:
                error = FieldValidationError('Описание не может быть длиннее 100 символов')
                error.field = 'description'
                raise error

            # Валидация даты создания
            if '.' in form_data['creation_date']:
                creation_date = validate_date_format(form_data['creation_date'])
            else:
                creation_date = datetime.strptime(form_data['creation_date'], '%Y-%m-%d').date()

            if creation_date > date.today():
                error = FieldValidationError('Дата создания не может быть в будущем!')
                error.field = 'creation_date'
                raise error

            # Валидация страны
            if form_data['country'] not in SPACE_COUNTRIES:
                error = FieldValidationError('Выберите страну из списка!')
                error.field = 'country'
                raise error

            # Валидация состояния
            if form_data['state'] not in EXHIBIT_STATES:
                error = FieldValidationError('Выберите состояние из списка!')
                error.field = 'state'
                raise error

            # Валидация типа
            if form_data['type'] not in EXHIBIT_TYPES:
                error = FieldValidationError('Выберите тип из списка!')
                error.field = 'type'
                raise error

            exhibit = Exhibit(
                exhibit_id=new_id,
                title=title,
                description=description,
                creation_date=creation_date,
                country=form_data['country'],
                state=form_data['state'],
                type=form_data['type']
            )

            # Обработка mission_id
            if form_data['mission_id'].strip():
                try:
                    mission = SpaceMission.objects.get(mission_id=int(form_data['mission_id']))
                    exhibit.mission_id = mission
                except (ValueError, SpaceMission.DoesNotExist):
                    error = FieldValidationError(f'Миссия с ID {form_data["mission_id"]} не найдена!')
                    error.field = 'mission_id'
                    raise error

            exhibit.save()
            messages.success(request, 'Экспонат успешно добавлен!')
            return redirect('exhibit_list')

        except ValidationError as e:
            messages.error(request, str(e))
            fields_data = {
                'countries': SPACE_COUNTRIES,
                'states': EXHIBIT_STATES,
                'types': EXHIBIT_TYPES,
                'form_data': form_data,
                'error_field': e.field if hasattr(e, 'field') else None
            }
            return render(request, 'core/exhibit_form_add.html', {
                'entity_name': 'Экспонат',
                'operation': 'Добавить',
                'entity_type': 'exhibit',
                'fields_data': fields_data
            })

    fields_data = {
        'countries': SPACE_COUNTRIES,
        'states': EXHIBIT_STATES,
        'types': EXHIBIT_TYPES,
        'form_data': form_data
    }

    return render(request, 'core/exhibit_form_add.html', {
        'entity_name': 'Экспонат',
        'operation': 'Добавить',
        'entity_type': 'exhibit',
        'fields_data': fields_data
    })


def exhibit_update(request):
    exhibits = Exhibit.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        exhibit = get_object_or_404(Exhibit, exhibit_id=selected_id)

        if request.method == 'POST':
            try:
                exhibit.title = request.POST['title']
                exhibit.description = request.POST['description']

                # Преобразуем строку в дату
                creation_date = datetime.strptime(request.POST['creation_date'], '%Y-%m-%d').date()

                # Валидация
                if creation_date > date.today():
                    messages.error(request, 'Дата создания не может быть в будущем!')
                    return redirect(f'exhibit_update?id={selected_id}')

                exhibit.creation_date = creation_date
                exhibit.country = request.POST['country']
                exhibit.state = request.POST['state']
                exhibit.type = request.POST['type']

                mission_id_value = request.POST.get('mission_id', '').strip()
                if mission_id_value:
                    try:
                        mission = SpaceMission.objects.get(mission_id=int(mission_id_value))
                        exhibit.mission_id = mission
                    except (ValueError, SpaceMission.DoesNotExist):
                        messages.error(request, f'Миссия с ID {mission_id_value} не найдена!')
                        return redirect(f'exhibit_update?id={selected_id}')
                else:
                    exhibit.mission_id = None

                exhibit.save()
                messages.success(request, 'Экспонат успешно обновлен!')
                return redirect('exhibit_list')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')

        fields = [
            {'name': 'title', 'label': 'Название', 'type': 'text', 'required': True, 'value': exhibit.title},
            {'name': 'description', 'label': 'Описание', 'type': 'textarea', 'required': True,
             'value': exhibit.description},
            {'name': 'creation_date', 'label': 'Дата создания', 'type': 'date', 'required': True,
             'value': exhibit.creation_date},
            {'name': 'country', 'label': 'Страна', 'type': 'text', 'required': True, 'value': exhibit.country},
            {'name': 'state', 'label': 'Состояние', 'type': 'text', 'required': True, 'value': exhibit.state},
            {'name': 'type', 'label': 'Тип', 'type': 'text', 'required': True, 'value': exhibit.type},
            {'name': 'mission_id', 'label': 'ID миссии (необязательно)', 'type': 'number', 'required': False,
             'value': exhibit.mission_id.mission_id if exhibit.mission_id else ''},
        ]

        return render(request, 'core/entity_form_update.html', {
            'entity_name': 'Экспонат',
            'operation': 'Обновить',
            'fields': fields,
            'entity_type': 'exhibit'
        })

    columns = ['exhibit_id', 'title', 'country', 'type']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Экспонат',
        'items': exhibits,
        'columns': columns,
        'entity_type': 'exhibit',
        'operation': 'обновления'
    })


def exhibit_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                exhibit = Exhibit.objects.get(exhibit_id=delete_id)
                exhibit.delete()
                messages.success(request, f'Экспонат "{exhibit.title}" успешно удален!')
            except Exhibit.DoesNotExist:
                messages.error(request, 'Экспонат не найден!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('exhibit_delete')

    exhibits = Exhibit.objects.all()
    columns = ['exhibit_id', 'title', 'country', 'type']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Экспонат',
        'items': exhibits,
        'columns': columns,
        'entity_type': 'exhibit',
        'operation': 'удаления'
    })


# ============= SPACE MISSION CRUD =============
def mission_list(request):
    missions = SpaceMission.objects.all()

    filter_column = request.GET.get('filter_column', '')
    filter_value = request.GET.get('filter_value', '')

    if filter_column and filter_value:
        if filter_column == 'crew':
            # Специальный фильтр для массива crew
            missions = missions.filter(crew__contains=[filter_value])
        else:
            filter_dict = {filter_column + '__icontains': filter_value}
            missions = missions.filter(**filter_dict)

    if request.GET.get('save_result'):
        # Преобразуем crew в строку для сохранения
        data = list(missions.values())
        for item in data:
            if 'crew' in item and item['crew']:
                item['crew'] = ', '.join(item['crew'])
        return save_query_result(data, 'missions_filtered')

    # Преобразуем crew в строку для отображения
    missions_display = []
    for mission in missions:
        mission_dict = {
            'mission_id': mission.mission_id,
            'title': mission.title,
            'country': mission.country,
            'start_date': mission.start_date,
            'end_date': mission.end_date,
            'crew': ', '.join(mission.crew) if mission.crew else '',
            'goal': mission.goal
        }
        missions_display.append(mission_dict)

    columns = ['mission_id', 'title', 'country', 'start_date', 'end_date', 'crew', 'goal']

    return render(request, 'core/entity_list.html', {
        'entity_name': 'Космические миссии',
        'items': missions_display,
        'columns': columns,
        'entity_type': 'mission',
        'filter_column': filter_column,
        'filter_value': filter_value
    })


def mission_add(request):
    form_data = {}  # Инициализируем пустой словарь

    if request.method == 'POST':
        form_data = {
            'title': request.POST.get('title', ''),
            'countries': request.POST.getlist('countries[]'),
            'start_date': request.POST.get('start_date', ''),
            'end_date': request.POST.get('end_date', ''),
            'crew': request.POST.getlist('crew[]'),
            'goal': request.POST.get('goal', ''),
        }

        try:
            # Получаем максимальный ID
            max_id = SpaceMission.objects.aggregate(models.Max('mission_id'))['mission_id__max']
            new_id = (max_id or 0) + 1

            # Проверка уникальности ID
            check_unique_id(SpaceMission, 'mission_id', new_id)

            # Валидация названия (буквы, цифры, -, скобки)
            title = form_data['title'].strip()
            if len(title) < 2:
                error = FieldValidationError('Название должно содержать минимум 2 символа')
                error.field = 'title'
                raise error
            if len(title) > 50:
                error = FieldValidationError('Название не может быть длиннее 50 символов')
                error.field = 'title'
                raise error
            if not re.match(r'^[а-яА-ЯёЁa-zA-Z0-9\s\(\)\-]+$', title):
                error = FieldValidationError('Название содержит недопустимые символы')
                error.field = 'title'
                raise error

            # Валидация стран
            if not form_data['countries']:
                error = FieldValidationError('Выберите хотя бы одну страну!')
                error.field = 'countries'
                raise error
            if len(form_data['countries']) > 5:
                error = FieldValidationError('Можно выбрать максимум 5 стран!')
                error.field = 'countries'
                raise error
            for country in form_data['countries']:
                if country not in SPACE_COUNTRIES:
                    error = FieldValidationError(f'Недопустимая страна: {country}')
                    error.field = 'countries'
                    raise error

            # Объединяем страны в строку
            country_str = ', '.join(form_data['countries'])

            # Валидация дат
            if '.' in form_data['start_date']:
                start_date = validate_date_format(form_data['start_date'])
            else:
                start_date = datetime.strptime(form_data['start_date'], '%Y-%m-%d').date()

            # Обработка даты окончания (может быть None)
            end_date = None
            if form_data['end_date'] and form_data['end_date'] != 'None':
                if '.' in form_data['end_date']:
                    end_date = validate_date_format(form_data['end_date'])
                else:
                    end_date = datetime.strptime(form_data['end_date'], '%Y-%m-%d').date()

                if end_date < start_date:
                    error = FieldValidationError('Дата окончания не может быть раньше даты начала!')
                    error.field = 'end_date'
                    raise error

            # Валидация экипажа
            crew_list = []
            for member in form_data['crew']:
                member = member.strip()
                if member:
                    # Валидация каждого члена экипажа как ФИО
                    try:
                        validated_member = validate_fio(member)
                        crew_list.append(validated_member)
                    except ValidationError:
                        error = FieldValidationError(f'Некорректное ФИО члена экипажа: {member}')
                        error.field = 'crew'
                        raise error

            # Валидация цели
            goal = form_data['goal'].strip()
            if len(goal) < 10:
                error = FieldValidationError('Цель миссии должна содержать минимум 10 символов')
                error.field = 'goal'
                raise error
            if len(goal) > 100:
                error = FieldValidationError('Цель миссии не может быть длиннее 100 символов')
                error.field = 'goal'
                raise error

            mission = SpaceMission(
                mission_id=new_id,
                title=title,
                country=country_str,
                start_date=start_date,
                end_date=end_date,
                crew=crew_list if crew_list else [],
                goal=goal
            )

            mission.save()
            messages.success(request, 'Космическая миссия успешно добавлена!')
            return redirect('mission_list')

        except ValidationError as e:
            messages.error(request, str(e))
            fields_data = {
                'countries': SPACE_COUNTRIES,
                'form_data': form_data,
                'error_field': e.field if hasattr(e, 'field') else None
            }
            return render(request, 'core/mission_form_add.html', {
                'entity_name': 'Космическая миссия',
                'operation': 'Добавить',
                'entity_type': 'mission',
                'fields_data': fields_data
            })

    fields_data = {
        'countries': SPACE_COUNTRIES,
        'form_data': form_data,
        'error_field': None
    }

    return render(request, 'core/mission_form_add.html', {
        'entity_name': 'Космическая миссия',
        'operation': 'Добавить',
        'entity_type': 'mission',
        'fields_data': fields_data
    })

def mission_update(request):
    missions = SpaceMission.objects.all()
    selected_id = request.GET.get('id')

    if selected_id:
        mission = get_object_or_404(SpaceMission, mission_id=selected_id)

        if request.method == 'POST':
            try:
                mission.title = request.POST['title']
                mission.country = request.POST['country']

                # Преобразуем строки в даты
                start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()
                end_date = datetime.strptime(request.POST['end_date'], '%Y-%m-%d').date()

                # Валидация
                if end_date < start_date:
                    messages.error(request, 'Дата окончания не может быть раньше даты начала!')
                    return redirect(f'mission_update?id={selected_id}')

                mission.start_date = start_date
                mission.end_date = end_date

                # Обработка экипажа
                crew_list = request.POST.getlist('crew[]')
                crew_list = [member.strip() for member in crew_list if member.strip()]
                mission.crew = crew_list if crew_list else []

                mission.goal = request.POST['goal']

                mission.save()
                messages.success(request, 'Космическая миссия успешно обновлена!')
                return redirect('mission_list')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')

        # Передаем текущий экипаж для отображения
        return render(request, 'core/mission_form_update.html', {
            'entity_name': 'Космическая миссия',
            'operation': 'Обновить',
            'entity_type': 'mission',
            'mission': mission,
            'crew_members': mission.crew if mission.crew else []
        })

    columns = ['mission_id', 'title', 'country', 'goal']
    return render(request, 'core/entity_select.html', {
        'entity_name': 'Космическая миссия',
        'items': missions,
        'columns': columns,
        'entity_type': 'mission',
        'operation': 'обновления'
    })


def mission_delete(request):
    if request.method == 'POST':
        delete_id = request.POST.get('delete_id')
        if delete_id:
            try:
                mission = SpaceMission.objects.get(mission_id=delete_id)
                mission.delete()
                messages.success(request, f'Миссия "{mission.title}" успешно удалена!')
            except SpaceMission.DoesNotExist:
                messages.error(request, 'Миссия не найдена!')
            except Exception as e:
                messages.error(request, f'Ошибка при удалении: {str(e)}')
            return redirect('mission_delete')

    missions = SpaceMission.objects.all()
    columns = ['mission_id', 'title', 'country', 'goal']

    return render(request, 'core/entity_select.html', {
        'entity_name': 'Космическая миссия',
        'items': missions,
        'columns': columns,
        'entity_type': 'mission',
        'operation': 'удаления'
    })


# ============= UTILITY FUNCTIONS =============
def save_query_result(data, name):
    """Сохранение результата запроса в Excel"""
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'{name}_{timestamp}.xlsx'
    filepath = os.path.join(backup_dir, filename)

    df = pd.DataFrame(data)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=name[:31])

        worksheet = writer.sheets[name[:31]]
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            ) + 2
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = min(max_length, 50)

    return HttpResponse(f'Результат запроса сохранен в файл {filename}')


def save_database_state(request):
    """Страница выбора формата сохранения БД"""
    stats = {}

    with connection.cursor() as cursor:
        # Получаем статистику
        cursor.execute("""
            SELECT COUNT(*) 
            FROM information_schema.tables 
            WHERE table_schema = 'public'
        """)
        stats['total_tables'] = cursor.fetchone()[0]

        cursor.execute("""
            SELECT COUNT(*) 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name NOT LIKE 'auth_%'
            AND table_name NOT LIKE 'django_%'
        """)
        stats['user_tables'] = cursor.fetchone()[0]

        # Подсчет общего количества записей в пользовательских таблицах
        total_records = 0
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name NOT LIKE 'auth_%'
            AND table_name NOT LIKE 'django_%'
        """)

        for row in cursor.fetchall():
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {row[0]}")
                total_records += cursor.fetchone()[0]
            except:
                pass

        stats['total_records'] = total_records

        # Размер БД
        cursor.execute("""
            SELECT pg_size_pretty(pg_database_size(current_database()))
        """)
        stats['db_size'] = cursor.fetchone()[0]

    return render(request, 'core/save_database.html', stats)


def save_database_excel(request):
    """Сохранение БД в Excel формат"""
    if request.method != 'POST':
        return redirect('save_database_state')

    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'database_backup_{timestamp}.xlsx'
    filepath = os.path.join(backup_dir, filename)

    try:
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            tables_saved = []

            with connection.cursor() as cursor:
                # Получаем список пользовательских таблиц
                cursor.execute("""
                    SELECT table_name 
                    FROM information_schema.tables 
                    WHERE table_schema = 'public' 
                    AND table_name NOT LIKE 'auth_%'
                    AND table_name NOT LIKE 'django_%'
                    ORDER BY table_name
                """)

                user_tables = [row[0] for row in cursor.fetchall()]

                # Сохраняем каждую таблицу
                for table_name in user_tables:
                    try:
                        # Получаем данные из таблицы
                        cursor.execute(f"SELECT * FROM {table_name}")
                        columns = [desc[0] for desc in cursor.description]
                        data = cursor.fetchall()

                        if data:
                            df = pd.DataFrame(data, columns=columns)

                            # Обработка специальных типов данных
                            for col in df.columns:
                                # Преобразуем списки/массивы в строки
                                if df[col].dtype == 'object':
                                    df[col] = df[col].apply(lambda x:
                                                            ', '.join(x) if isinstance(x, list) else x
                                                            )

                            # Сохраняем на отдельный лист (максимум 31 символ для имени листа)
                            sheet_name = table_name[:31]
                            df.to_excel(writer, sheet_name=sheet_name, index=False)

                            # Настройка ширины колонок
                            worksheet = writer.sheets[sheet_name]
                            for idx, col in enumerate(df.columns):
                                max_length = max(
                                    df[col].astype(str).map(len).max() if not df.empty else 0,
                                    len(str(col))
                                ) + 2
                                worksheet.column_dimensions[
                                    openpyxl.utils.get_column_letter(idx + 1)
                                ].width = min(max_length, 50)

                            tables_saved.append(table_name)
                        else:
                            # Если таблица пустая, создаем лист с заголовками
                            df = pd.DataFrame(columns=columns)
                            sheet_name = table_name[:31]
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            tables_saved.append(f"{table_name} (пустая)")

                    except Exception as e:
                        print(f"Ошибка при сохранении таблицы {table_name}: {e}")

            # Добавляем информационный лист
            info_data = {
                'Информация': ['Дата создания', 'Количество таблиц', 'Сохраненные таблицы'],
                'Значение': [
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    len(tables_saved),
                    ', '.join(tables_saved)
                ]
            }
            info_df = pd.DataFrame(info_data)
            info_df.to_excel(writer, sheet_name='INFO', index=False)

        # Отправляем файл пользователю
        with open(filepath, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

    except Exception as e:
        messages.error(request, f'Ошибка при создании Excel файла: {str(e)}')
        return redirect('save_database_state')


def save_database_sql(request):
    """Сохранение БД в SQL формат"""
    if request.method != 'POST':
        return redirect('save_database_state')

    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    filename = f'database_dump_{timestamp}.sql'
    filepath = os.path.join(backup_dir, filename)

    try:
        with open(filepath, 'w', encoding='utf-8') as sql_file:
            # Заголовок файла
            sql_file.write("-- Space Museum Database Dump\n")
            sql_file.write(f"-- Generated: {datetime.now()}\n")
            sql_file.write("-- Database: PostgreSQL\n\n")
            sql_file.write("SET client_encoding = 'UTF8';\n\n")

            with connection.cursor() as cursor:
                # Получаем список пользовательских таблиц
                cursor.execute("""
                    SELECT table_name 
                    FROM information_schema.tables 
                    WHERE table_schema = 'public' 
                    AND table_name NOT LIKE 'auth_%'
                    AND table_name NOT LIKE 'django_%'
                    ORDER BY table_name
                """)

                user_tables = [row[0] for row in cursor.fetchall()]

                for table_name in user_tables:
                    sql_file.write(f"\n-- Table: {table_name}\n")
                    sql_file.write(f"-- DROP TABLE IF EXISTS {table_name} CASCADE;\n\n")

                    # Получаем структуру таблицы
                    sql_file.write(f"CREATE TABLE IF NOT EXISTS {table_name} (\n")

                    cursor.execute("""
                        SELECT 
                            column_name,
                            data_type,
                            character_maximum_length,
                            is_nullable,
                            column_default
                        FROM information_schema.columns
                        WHERE table_name = %s
                        ORDER BY ordinal_position
                    """, [table_name])

                    columns = cursor.fetchall()
                    column_definitions = []

                    for col in columns:
                        col_def = f"    {col[0]} "

                        # Тип данных
                        if col[1] == 'character varying' and col[2]:
                            col_def += f"VARCHAR({col[2]})"
                        elif col[1] == 'integer':
                            col_def += "INTEGER"
                        elif col[1] == 'bigint':
                            col_def += "BIGINT"
                        elif col[1] == 'text':
                            col_def += "TEXT"
                        elif col[1] == 'date':
                            col_def += "DATE"
                        elif col[1] == 'timestamp without time zone':
                            col_def += "TIMESTAMP"
                        elif col[1] == 'boolean':
                            col_def += "BOOLEAN"
                        elif col[1] == 'ARRAY':
                            col_def += "TEXT[]"
                        else:
                            col_def += col[1].upper()

                        # NOT NULL
                        if col[3] == 'NO':
                            col_def += " NOT NULL"

                        # DEFAULT
                        if col[4]:
                            col_def += f" DEFAULT {col[4]}"

                        column_definitions.append(col_def)

                    # Получаем PRIMARY KEY
                    cursor.execute("""
                        SELECT kcu.column_name
                        FROM information_schema.table_constraints tc
                        JOIN information_schema.key_column_usage kcu
                            ON tc.constraint_name = kcu.constraint_name
                        WHERE tc.table_name = %s 
                        AND tc.constraint_type = 'PRIMARY KEY'
                    """, [table_name])

                    pk_columns = [row[0] for row in cursor.fetchall()]
                    if pk_columns:
                        column_definitions.append(
                            f"    PRIMARY KEY ({', '.join(pk_columns)})"
                        )

                    # Получаем FOREIGN KEYS
                    cursor.execute("""
                        SELECT
                            kcu.column_name,
                            ccu.table_name AS foreign_table_name,
                            ccu.column_name AS foreign_column_name,
                            rc.delete_rule
                        FROM information_schema.table_constraints AS tc
                        JOIN information_schema.key_column_usage AS kcu
                            ON tc.constraint_name = kcu.constraint_name
                        JOIN information_schema.constraint_column_usage AS ccu
                            ON ccu.constraint_name = tc.constraint_name
                        JOIN information_schema.referential_constraints rc
                            ON tc.constraint_name = rc.constraint_name
                        WHERE tc.constraint_type = 'FOREIGN KEY' 
                        AND tc.table_name = %s
                    """, [table_name])

                    for fk in cursor.fetchall():
                        fk_def = f"    FOREIGN KEY ({fk[0]}) REFERENCES {fk[1]}({fk[2]})"
                        if fk[3]:
                            fk_def += f" ON DELETE {fk[3]}"
                        column_definitions.append(fk_def)

                    sql_file.write(',\n'.join(column_definitions))
                    sql_file.write("\n);\n\n")

                    # Получаем данные таблицы
                    cursor.execute(f"SELECT * FROM {table_name}")
                    rows = cursor.fetchall()

                    if rows:
                        columns = [desc[0] for desc in cursor.description]
                        sql_file.write(f"-- Data for {table_name}\n")

                        for row in rows:
                            values = []
                            for val in row:
                                if val is None:
                                    values.append("NULL")
                                elif isinstance(val, bool):
                                    values.append("TRUE" if val else "FALSE")
                                elif isinstance(val, (int, float)):
                                    values.append(str(val))
                                elif isinstance(val, list):
                                    # PostgreSQL array format
                                    array_vals = [f"'{v}'" for v in val]
                                    values.append("ARRAY[" + ", ".join(array_vals) + "]")
                                elif isinstance(val, (date, datetime)):
                                    values.append(f"'{val}'")
                                else:
                                    # Экранируем одинарные кавычки
                                    escaped_val = str(val).replace("'", "''")
                                    values.append(f"'{escaped_val}'")

                            insert_sql = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(values)});\n"
                            sql_file.write(insert_sql)

                        sql_file.write("\n")

                # Добавляем команды для последовательностей (sequences)
                sql_file.write("\n-- Reset sequences\n")
                for table_name in user_tables:
                    cursor.execute("""
                        SELECT column_name
                        FROM information_schema.columns
                        WHERE table_name = %s
                        AND column_default LIKE 'nextval%%'
                    """, [table_name])

                    for col in cursor.fetchall():
                        sql_file.write(
                            f"SELECT setval(pg_get_serial_sequence('{table_name}', '{col[0]}'), "
                            f"COALESCE((SELECT MAX({col[0]}) FROM {table_name}), 1));\n"
                        )

        # Отправляем файл пользователю
        with open(filepath, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/sql'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

    except Exception as e:
        messages.error(request, f'Ошибка при создании SQL дампа: {str(e)}')
        return redirect('save_database_state')

def special_queries(request):
    """Страница для специальных SQL запросов"""
    query_result = None
    columns = []
    lab_type = request.POST.get('lab_type', '')
    selected_query = request.POST.get('query_type', '')

    # Словарь всех запросов Лабораторной работы 5
    lab5_queries = {
        # Employee запросы
        'emp_1': """SELECT e.full_name AS "ФИО", e.position FROM employee e""",
        'emp_2': """SELECT * FROM employee WHERE position = 'Экскурсовод'""",
        'emp_3': """SELECT * FROM employee ORDER BY hire_date ASC""",
        'emp_4': """SELECT * FROM employee WHERE full_name LIKE 'П%'""",

        # Excursion запросы
        'exc_1': """SELECT e.title AS "Название выставки", e.date FROM excursion e""",
        'exc_2': """SELECT title, duration FROM excursion ORDER BY duration DESC""",
        'exc_3': """SELECT DISTINCT title FROM excursion WHERE title LIKE '%Косм%'""",
        'exc_4': """SELECT e.full_name, ex.title AS excursion 
                    FROM employee AS e 
                    INNER JOIN excursion AS ex ON e.employee_id = ex.employee_id""",
        'exc_5': """SELECT e.full_name, e.position, ex.title AS excursion_title, ex.language 
                    FROM employee AS e 
                    LEFT JOIN excursion AS ex ON e.employee_id = ex.employee_id""",
        'exc_6': """SELECT ex.title AS excursion_title, ex.language, e.full_name, e.position 
                    FROM employee AS e 
                    RIGHT JOIN excursion AS ex ON e.employee_id = ex.employee_id""",
        'exc_7': """SELECT e.employee_id, e.full_name, e.position, 
                    ex.excursion_id, ex.title AS excursion_title, ex.date, ex.language 
                    FROM employee AS e 
                    FULL JOIN excursion AS ex ON e.employee_id = ex.employee_id""",

        # Exhibit запросы
        'exh_1': """SELECT e.title AS "Название экспоната", e.state FROM exhibit e""",
        'exh_2': """SELECT title, creation_date, country FROM exhibit 
                    WHERE creation_date < '1970-01-01' AND country = 'СССР'""",
        'exh_3': """SELECT title, creation_date FROM exhibit ORDER BY creation_date ASC""",
        'exh_4': """SELECT DISTINCT title FROM exhibit WHERE title LIKE '%Макет%'""",
        'exh_5': """SELECT e.title AS exhibit_title, e.type, m.title AS mission_name, m.start_date 
                    FROM exhibit AS e 
                    INNER JOIN space_mission AS m ON e.mission_id = m.mission_id""",
        'exh_6': """SELECT e.title AS exhibit_title, e.country, e.type, 
                    m.title AS mission_name, m.start_date 
                    FROM exhibit AS e 
                    LEFT JOIN space_mission AS m ON e.mission_id = m.mission_id""",
        'exh_7': """SELECT m.title AS mission_name, m.start_date, 
                    e.title AS exhibit_title, e.type 
                    FROM exhibit AS e 
                    RIGHT JOIN space_mission AS m ON e.mission_id = m.mission_id""",
        'exh_8': """SELECT e.exhibit_id, e.title AS exhibit_title, e.type, 
                    m.mission_id, m.title AS mission_name, m.start_date 
                    FROM exhibit AS e 
                    FULL JOIN space_mission AS m ON e.mission_id = m.mission_id""",

        # Space Mission запросы
        'sm_1': """SELECT s.title AS "Название Миссии", s.country FROM space_mission s""",
        'sm_2': """SELECT title, start_date, end_date FROM space_mission 
                   WHERE start_date > '1970-01-01' AND end_date < '2000-01-01'""",
        'sm_3': """SELECT title, country FROM space_mission ORDER BY country DESC""",
        'sm_4': """SELECT title, country FROM space_mission WHERE country LIKE '%США%'""",

        # Exhibition запросы
        'exb_1': """SELECT e.title AS "Название выставки", e.theme FROM exhibition e""",
        'exb_2': """SELECT title, start_date, end_date, type FROM exhibition 
                    WHERE type = 'Временная' AND end_date IS NOT NULL""",
        'exb_3': """SELECT title, start_date, end_date FROM exhibition ORDER BY end_date ASC""",
        'exb_4': """SELECT title, location FROM exhibition WHERE location LIKE '%Зал%'""",
        'exb_5': """SELECT ex.title AS exhibition_title, ex.location, e.full_name, e.position 
                    FROM exhibition AS ex 
                    INNER JOIN exhibition_employee AS ee ON ex.exhibition_id = ee.exhibition_id 
                    INNER JOIN employee AS e ON ee.employee_id = e.employee_id""",
        'exb_6': """SELECT ex.title AS exhibition_title, ex.theme, ex.location, 
                    e.full_name, e.position 
                    FROM exhibition AS ex 
                    LEFT JOIN exhibition_employee AS ee ON ex.exhibition_id = ee.exhibition_id 
                    LEFT JOIN employee AS e ON ee.employee_id = e.employee_id""",
        'exb_7': """SELECT e.full_name, e.position, ex.title AS exhibition_title, ex.location 
                    FROM exhibition AS ex 
                    RIGHT JOIN exhibition_employee AS ee ON ex.exhibition_id = ee.exhibition_id 
                    RIGHT JOIN employee AS e ON ee.employee_id = e.employee_id""",
        'exb_8': """SELECT ex.exhibition_id, ex.title AS exhibition_title, ex.location, 
                    e.employee_id, e.full_name, e.position 
                    FROM exhibition AS ex 
                    FULL JOIN exhibition_employee AS ee ON ex.exhibition_id = ee.exhibition_id 
                    FULL JOIN employee AS e ON ee.employee_id = e.employee_id""",

        # Visitor запросы
        'vis_1': """SELECT v.ticket_type AS "Тип билета", v.full_name, v.citizenship FROM visitor v""",
        'vis_2': """SELECT full_name, review FROM visitor WHERE review IS NOT NULL""",
        'vis_3': """SELECT full_name, birth_date FROM visitor ORDER BY birth_date DESC""",
        'vis_4': """SELECT full_name, review FROM visitor WHERE review LIKE '%!%'""",
        'vis_5': """SELECT ex.title AS excursion_title, ex.date AS excursion_date, 
                    v.full_name AS visitor_name, v.ticket_type 
                    FROM excursion AS ex 
                    INNER JOIN excursion_visitor AS ev ON ex.excursion_id = ev.excursion_id 
                    INNER JOIN visitor AS v ON ev.visitor_id = v.visitor_id""",
        'vis_6': """SELECT ex.excursion_id, ex.title AS excursion_title, ex.date, ex.language, 
                    v.visitor_id, v.full_name AS visitor_name, v.visit_date 
                    FROM excursion AS ex 
                    LEFT JOIN excursion_visitor AS ev ON ex.excursion_id = ev.excursion_id 
                    LEFT JOIN visitor AS v ON ev.visitor_id = v.visitor_id""",
        'vis_7': """SELECT v.full_name AS visitor_name, v.ticket_type, 
                    ex.excursion_id, ex.title AS excursion_title, ex.date 
                    FROM excursion AS ex 
                    RIGHT JOIN excursion_visitor AS ev ON ex.excursion_id = ev.excursion_id 
                    RIGHT JOIN visitor AS v ON ev.visitor_id = v.visitor_id""",
        'vis_8': """SELECT ex.excursion_id, ex.title AS excursion_title, ex.date, 
                    v.visitor_id, v.full_name AS visitor_name, v.visit_date 
                    FROM excursion AS ex 
                    FULL JOIN excursion_visitor AS ev ON ex.excursion_id = ev.excursion_id 
                    FULL JOIN visitor AS v ON ev.visitor_id = v.visitor_id"""
    }

    # Запросы для Лабораторной работы 6
    lab6_queries = {
        'lab6_1': """
            SELECT 
                e.employee_id,
                e.full_name,
                COUNT(ex.excursion_id) AS excursions_conducted,
                SUM(ex.ticket_num) AS total_tickets
            FROM employee e
            LEFT JOIN excursion ex ON e.employee_id = ex.employee_id
            WHERE e.department = 'Экскурсионный отдел'
            GROUP BY e.employee_id, e.full_name
            HAVING COUNT(ex.excursion_id) > ALL(
                SELECT COUNT(ex2.excursion_id)
                FROM employee e2
                LEFT JOIN excursion ex2 ON e2.employee_id = ex2.employee_id
                WHERE e2.department = 'Администрация'
                GROUP BY e2.employee_id
            )
            INTERSECT ALL
            SELECT 
                e.employee_id,
                e.full_name,
                COUNT(ex.excursion_id) AS excursions_conducted,
                SUM(ex.ticket_num) AS total_tickets
            FROM employee e
            LEFT JOIN excursion ex ON e.employee_id = ex.employee_id
            WHERE e.department = 'Экскурсионный отдел'
            GROUP BY e.employee_id, e.full_name
            HAVING COUNT(ex.excursion_id) >= 2 AND SUM(ex.ticket_num) > 100
            ORDER BY excursions_conducted DESC, total_tickets DESC
        """,

        'lab6_2': """
            SELECT excursion_id, title, language, duration, ticket_num
            FROM excursion
            WHERE language = 'Русский'
            GROUP BY excursion_id, title, language, duration, ticket_num
            HAVING duration = (SELECT MIN(duration) FROM excursion WHERE language = 'Русский')
            INTERSECT
            SELECT excursion_id, title, language, duration, ticket_num
            FROM excursion
            WHERE language = 'Русский'
            GROUP BY excursion_id, title, language, duration, ticket_num
            HAVING ticket_num > (SELECT AVG(ticket_num) FROM excursion WHERE language = 'Русский')
            ORDER BY duration, ticket_num
        """,

        'lab6_3': """
            WITH country_stats AS (
                SELECT 
                    country,
                    COUNT(*) as exhibit_count,
                    AVG(EXTRACT(YEAR FROM creation_date)) as avg_year
                FROM exhibit
                WHERE EXTRACT(YEAR FROM creation_date) > 1960
                GROUP BY country
                HAVING COUNT(*) > 2
            )
            SELECT country, exhibit_count, avg_year
            FROM country_stats
            UNION
            SELECT 
                country,
                COUNT(*) as exhibit_count,
                AVG(EXTRACT(YEAR FROM creation_date)) as avg_year
            FROM exhibit
            WHERE country NOT IN (
                SELECT DISTINCT country
                FROM exhibit
                WHERE type = 'Макет'
            )
            GROUP BY country
            ORDER BY exhibit_count DESC, avg_year DESC
        """,

        'lab6_4': """
            SELECT 
                country,
                COUNT(*) as mission_count,
                MAX(EXTRACT(YEAR FROM start_date)) as max_start_year
            FROM space_mission
            GROUP BY country
            HAVING COUNT(*) > ANY(
                SELECT COUNT(*)
                FROM space_mission
                GROUP BY country
                HAVING COUNT(*) < 3
            )
            UNION ALL
            SELECT 
                country,
                COUNT(*) as mission_count,
                MAX(EXTRACT(YEAR FROM start_date)) as max_start_year
            FROM space_mission
            WHERE goal LIKE '%первый%'
            GROUP BY country
            ORDER BY mission_count DESC, max_start_year DESC
        """,

        'lab6_5': """
            SELECT 
                e.exhibition_id,
                e.title,
                e.type,
                COUNT(ex.exhibit_id) as exhibit_count
            FROM exhibition e
            INNER JOIN exhibit ex ON e.exhibition_id = ex.exhibition_id
            WHERE e.location = 'Главный зал'
            GROUP BY e.exhibition_id, e.title, e.type
            HAVING COUNT(ex.exhibit_id) > 2
            EXCEPT ALL
            SELECT 
                e.exhibition_id,
                e.title,
                e.type,
                COUNT(ex.exhibit_id) as exhibit_count
            FROM exhibition e
            INNER JOIN exhibit ex ON e.exhibition_id = ex.exhibition_id
            WHERE e.type = 'Временная'
            GROUP BY e.exhibition_id, e.title, e.type
            ORDER BY exhibit_count DESC
        """,

        'lab6_6': """
            SELECT 
                visitor_id,
                full_name,
                birth_date,
                COUNT(review) as review_count
            FROM visitor
            WHERE review IS NOT NULL
            GROUP BY visitor_id, full_name, birth_date
            HAVING COUNT(review) >= 1
            EXCEPT
            SELECT 
                visitor_id,
                full_name,
                birth_date,
                COUNT(review) as review_count
            FROM visitor
            WHERE birth_date < '1950-01-01'
            GROUP BY visitor_id, full_name, birth_date
            ORDER BY birth_date DESC
        """
    }

    if request.method == 'POST':
        query_sql = None

        if lab_type == 'lab5' and selected_query in lab5_queries:
            query_sql = lab5_queries[selected_query]
        elif lab_type == 'lab6' and selected_query in lab6_queries:
            query_sql = lab6_queries[selected_query]

        if query_sql:
            try:
                with connection.cursor() as cursor:
                    cursor.execute(query_sql)
                    columns = [col[0] for col in cursor.description]
                    query_result = cursor.fetchall()
            except Exception as e:
                messages.error(request, f'Ошибка выполнения запроса: {str(e)}')

        # Сохранение результата
        if request.POST.get('save_result') and query_result:
            return save_query_result(
                [dict(zip(columns, row)) for row in query_result],
                'special_query'
            )

    return render(request, 'core/special_queries.html', {
        'query_result': query_result,
        'columns': columns,
        'selected_query': selected_query,
        'lab_type': lab_type
    })


def table_management(request):
    """Главная страница управления таблицами"""
    return render(request, 'core/table_management.html')


def table_add(request):
    """Создание новой таблицы"""
    if request.method == 'POST':
        try:
            table_name = request.POST.get('table_name')

            # Валидация имени таблицы
            if not table_name or not table_name.replace('_', '').isalnum():
                raise ValueError("Некорректное имя таблицы")

            # Построение SQL запроса
            sql_parts = [f"CREATE TABLE {table_name} ("]
            columns = []
            primary_keys = []

            # Обработка колонок
            for key in request.POST:
                if key.startswith('column_name_'):
                    idx = key.split('_')[-1]
                    col_name = request.POST.get(f'column_name_{idx}')
                    col_type = request.POST.get(f'column_type_{idx}')

                    if col_name and col_type:
                        col_def = f"{col_name} {col_type}"

                        # Проверка ограничений
                        if request.POST.get(f'column_primary_{idx}'):
                            primary_keys.append(col_name)

                        if request.POST.get(f'column_notnull_{idx}'):
                            col_def += " NOT NULL"

                        if request.POST.get(f'column_unique_{idx}'):
                            col_def += " UNIQUE"

                        default_val = request.POST.get(f'column_default_{idx}')
                        if default_val and default_val != 'NULL':
                            col_def += f" DEFAULT {default_val}"

                        columns.append(col_def)

            # Добавление primary key
            if primary_keys:
                columns.append(f"PRIMARY KEY ({', '.join(primary_keys)})")

            # Обработка foreign keys
            for key in request.POST:
                if key.startswith('fk_column_'):
                    idx = key.split('_')[-1]
                    fk_column = request.POST.get(f'fk_column_{idx}')
                    fk_ref_table = request.POST.get(f'fk_ref_table_{idx}')
                    fk_ref_column = request.POST.get(f'fk_ref_column_{idx}')
                    fk_on_delete = request.POST.get(f'fk_on_delete_{idx}', 'CASCADE')

                    if fk_column and fk_ref_table and fk_ref_column:
                        fk_def = f"FOREIGN KEY ({fk_column}) REFERENCES {fk_ref_table}({fk_ref_column}) ON DELETE {fk_on_delete}"
                        columns.append(fk_def)

            sql_parts.append(', '.join(columns))
            sql_parts.append(')')

            create_sql = ' '.join(sql_parts)

            # Выполнение SQL
            with connection.cursor() as cursor:
                cursor.execute(create_sql)

            messages.success(request, f'Таблица "{table_name}" успешно создана!')
            return redirect('table_management')

        except Exception as e:
            messages.error(request, f'Ошибка при создании таблицы: {str(e)}')

    # Получение списка существующих таблиц
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            ORDER BY table_name
        """)
        existing_tables = [row[0] for row in cursor.fetchall()]

    return render(request, 'core/table_add.html', {
        'existing_tables': json.dumps(existing_tables)
    })


def table_list(request):
    """Список таблиц для удаления"""
    if request.method == 'POST':
        table_name = request.POST.get('table_name')
        try:
            # Проверка защищенных таблиц
            protected_tables = ['auth_user', 'auth_group', 'django_migrations', 'django_content_type']
            if table_name in protected_tables:
                raise ValueError("Эта таблица защищена от удаления")

            with connection.cursor() as cursor:
                cursor.execute(f"DROP TABLE IF EXISTS {table_name} CASCADE")

            messages.success(request, f'Таблица "{table_name}" успешно удалена!')

        except Exception as e:
            messages.error(request, f'Ошибка при удалении таблицы: {str(e)}')

        return redirect('table_list')

    # Получение информации о таблицах
    tables = []
    with connection.cursor() as cursor:
        # Получаем все таблицы
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name NOT LIKE 'django_%'
            AND table_name NOT LIKE 'auth_%'
            ORDER BY table_name
        """)

        for row in cursor.fetchall():
            table_name = row[0]

            # Получаем количество записей
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                row_count = cursor.fetchone()[0]
            except:
                row_count = 0

            # Получаем зависимости
            cursor.execute("""
                SELECT DISTINCT
                    tc.table_name AS referencing_table
                FROM information_schema.table_constraints AS tc
                JOIN information_schema.key_column_usage AS kcu
                    ON tc.constraint_name = kcu.constraint_name
                JOIN information_schema.constraint_column_usage AS ccu
                    ON ccu.constraint_name = tc.constraint_name
                WHERE tc.constraint_type = 'FOREIGN KEY'
                    AND ccu.table_name = %s
            """, [table_name])

            dependencies = [dep[0] for dep in cursor.fetchall()]

            tables.append({
                'name': table_name,
                'row_count': row_count,
                'dependencies': dependencies,
                'has_dependencies': len(dependencies) > 0
            })

    return render(request, 'core/table_list.html', {'tables': tables})


def table_info(request):
    """Информация о структуре БД (только пользовательские таблицы)"""
    tables_info = []

    with connection.cursor() as cursor:
        # Получаем только пользовательские таблицы
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name NOT LIKE 'auth_%'
            AND table_name NOT LIKE 'django_%'
            ORDER BY table_name
        """)

        for row in cursor.fetchall():
            table_name = row[0]
            table_data = {'name': table_name, 'columns': [], 'foreign_keys': []}

            # Получаем количество записей
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                table_data['row_count'] = cursor.fetchone()[0]
            except:
                table_data['row_count'] = 0

            # Получаем информацию о колонках
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    column_default,
                    character_maximum_length
                FROM information_schema.columns
                WHERE table_name = %s
                ORDER BY ordinal_position
            """, [table_name])

            for col in cursor.fetchall():
                col_info = {
                    'name': col[0],
                    'type': col[1],
                    'null': col[2] == 'YES',
                    'default': col[3],
                    'key': ''
                }

                # Определяем тип ключа
                cursor.execute("""
                    SELECT constraint_type
                    FROM information_schema.key_column_usage kcu
                    JOIN information_schema.table_constraints tc
                        ON kcu.constraint_name = tc.constraint_name
                    WHERE kcu.table_name = %s AND kcu.column_name = %s
                """, [table_name, col[0]])

                key_result = cursor.fetchone()
                if key_result:
                    if 'PRIMARY' in key_result[0]:
                        col_info['key'] = 'PRI'
                    elif 'FOREIGN' in key_result[0]:
                        col_info['key'] = 'MUL'

                table_data['columns'].append(col_info)

            # Получаем foreign keys
            cursor.execute("""
                SELECT
                    kcu.column_name,
                    ccu.table_name AS foreign_table_name,
                    ccu.column_name AS foreign_column_name
                FROM information_schema.table_constraints AS tc
                JOIN information_schema.key_column_usage AS kcu
                    ON tc.constraint_name = kcu.constraint_name
                JOIN information_schema.constraint_column_usage AS ccu
                    ON ccu.constraint_name = tc.constraint_name
                WHERE tc.constraint_type = 'FOREIGN KEY' AND tc.table_name = %s
            """, [table_name])

            for fk in cursor.fetchall():
                table_data['foreign_keys'].append({
                    'column': fk[0],
                    'ref_table': fk[1],
                    'ref_column': fk[2]
                })

            tables_info.append(table_data)

    return render(request, 'core/table_info.html', {'tables_info': tables_info})


def get_table_columns(request):
    """AJAX endpoint для получения колонок таблицы"""
    table_name = request.GET.get('table')
    columns = []

    if table_name:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = %s
                ORDER BY ordinal_position
            """, [table_name])

            columns = [row[0] for row in cursor.fetchall()]

    return JsonResponse({'columns': columns})


def table_columns_manage(request):
    """Управление колонками таблицы"""
    table_name = request.GET.get('table') or request.POST.get('table')

    if not table_name:
        return redirect('universal_crud')

    if request.method == 'POST':
        action = request.POST.get('action')

        with connection.cursor() as cursor:
            if action == 'add_column':
                # Добавление колонки
                column_name = request.POST.get('new_column_name')
                column_type = request.POST.get('new_column_type')
                nullable = request.POST.get('new_column_nullable')
                default = request.POST.get('new_column_default')

                sql = f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"

                if not nullable:
                    sql += " NOT NULL"

                if default and default != 'NULL':
                    sql += f" DEFAULT {default}"

                try:
                    cursor.execute(sql)
                    messages.success(request, f'Колонка {column_name} успешно добавлена')
                except Exception as e:
                    messages.error(request, f'Ошибка при добавлении колонки: {str(e)}')

            elif action == 'drop_column':
                # Удаление колонки
                column_name = request.POST.get('column_name')

                try:
                    cursor.execute(f"ALTER TABLE {table_name} DROP COLUMN {column_name}")
                    messages.success(request, f'Колонка {column_name} успешно удалена')
                except Exception as e:
                    messages.error(request, f'Ошибка при удалении колонки: {str(e)}')

        return redirect(f"{request.path}?table={table_name}")

    # Получаем информацию о колонках
    columns = []
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                c.column_name,
                c.data_type,
                c.is_nullable,
                c.column_default,
                CASE WHEN tc.constraint_type = 'PRIMARY KEY' THEN true ELSE false END as is_primary
            FROM information_schema.columns c
            LEFT JOIN information_schema.key_column_usage kcu
                ON c.table_name = kcu.table_name AND c.column_name = kcu.column_name
            LEFT JOIN information_schema.table_constraints tc
                ON kcu.constraint_name = tc.constraint_name AND tc.constraint_type = 'PRIMARY KEY'
            WHERE c.table_name = %s
            ORDER BY c.ordinal_position
        """, [table_name])

        for row in cursor.fetchall():
            columns.append({
                'name': row[0],
                'type': row[1],
                'nullable': row[2] == 'YES',
                'default': row[3],
                'is_primary': row[4]
            })

    return render(request, 'core/table_columns_manage.html', {
        'table_name': table_name,
        'columns': columns
    })


def table_restore_excel(request):
    """Восстановление таблицы из Excel файла с улучшенным определением типов"""
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        custom_table_name = request.POST.get('table_name', '').strip()
        replace_data = request.POST.get('replace_data')

        if excel_file:
            try:
                # Читаем Excel файл
                df = pd.read_excel(excel_file)

                # Определяем имя таблицы
                if custom_table_name:
                    table_name = custom_table_name
                else:
                    # Используем имя листа или файла
                    xls = pd.ExcelFile(excel_file)
                    table_name = xls.sheet_names[0].lower().replace(' ', '_')

                # Проверяем корректность имени
                if not table_name.replace('_', '').isalnum():
                    raise ValueError("Некорректное имя таблицы")

                with connection.cursor() as cursor:
                    # Проверяем существование таблицы
                    cursor.execute("""
                        SELECT EXISTS (
                            SELECT 1 FROM information_schema.tables 
                            WHERE table_schema = 'public' 
                            AND table_name = %s
                        )
                    """, [table_name])

                    table_exists = cursor.fetchone()[0]

                    if not table_exists:
                        # Улучшенное определение типов данных
                        columns_sql = []
                        column_types = {}

                        for col in df.columns:
                            # Очищаем имя колонки
                            col_name = col.lower().replace(' ', '_').replace('-', '_')

                            # Определяем тип данных более точно
                            sample = df[col].dropna()
                            if sample.empty:
                                col_type = "TEXT"
                            else:
                                # Проверяем на boolean
                                unique_values = sample.unique()
                                if len(unique_values) <= 2 and all(
                                        str(v).lower() in ['true', 'false', '1', '0', 'yes', 'no', 'да', 'нет', '1.0',
                                                           '0.0']
                                        or isinstance(v, bool)
                                        for v in unique_values
                                ):
                                    col_type = "BOOLEAN"
                                # Проверяем на integer
                                elif sample.dtype in ['int64', 'int32']:
                                    if sample.max() > 2147483647:
                                        col_type = "BIGINT"
                                    else:
                                        col_type = "INTEGER"
                                # Проверяем на float
                                elif sample.dtype in ['float64', 'float32']:
                                    # Проверяем, не являются ли float на самом деле integer
                                    if all(sample == sample.astype(int)):
                                        col_type = "INTEGER"
                                    else:
                                        col_type = "DECIMAL(10,2)"
                                # Проверяем на дату
                                else:
                                    try:
                                        pd.to_datetime(sample)
                                        # Проверяем, есть ли время
                                        if any(pd.to_datetime(sample).dt.time != pd.Timestamp('00:00:00').time()):
                                            col_type = "TIMESTAMP"
                                        else:
                                            col_type = "DATE"
                                    except:
                                        # Проверяем длину строк
                                        max_len = sample.astype(str).str.len().max()
                                        if max_len <= 255:
                                            col_type = f"VARCHAR({min(255, max_len * 2)})"
                                        else:
                                            col_type = "TEXT"

                            column_types[col_name] = col_type
                            columns_sql.append(f"{col_name} {col_type}")

                        # Добавляем ID если его нет
                        if 'id' not in [c.lower() for c in df.columns]:
                            columns_sql.insert(0, "id SERIAL PRIMARY KEY")
                        else:
                            # Если ID есть, делаем его PRIMARY KEY
                            for i, col_sql in enumerate(columns_sql):
                                if col_sql.startswith('id '):
                                    columns_sql[i] = col_sql.replace('INTEGER', 'INTEGER PRIMARY KEY').replace('BIGINT',
                                                                                                               'BIGINT PRIMARY KEY')
                                    break

                        create_sql = f"CREATE TABLE {table_name} ({', '.join(columns_sql)})"
                        cursor.execute(create_sql)

                        messages.info(request,
                                      f'Таблица {table_name} создана с колонками: {", ".join(column_types.keys())}')

                    elif replace_data:
                        # Очищаем существующую таблицу
                        cursor.execute(f"TRUNCATE TABLE {table_name} CASCADE")

                    # Вставляем данные с преобразованием типов
                    inserted_count = 0
                    for _, row in df.iterrows():
                        columns = []
                        values = []

                        # Если таблица была создана с ID, генерируем его
                        if not table_exists and 'id' not in [c.lower() for c in df.columns]:
                            cursor.execute(f"SELECT COALESCE(MAX(id), 0) + 1 FROM {table_name}")
                            new_id = cursor.fetchone()[0]
                            columns.append('id')
                            values.append(new_id)

                        for col in df.columns:
                            col_name = col.lower().replace(' ', '_').replace('-', '_')
                            value = row[col]

                            if pd.notna(value):
                                columns.append(col_name)

                                # Преобразуем boolean значения
                                if not table_exists and column_types.get(col_name) == 'BOOLEAN':
                                    if str(value).lower() in ['true', '1', 'yes', 'да', '1.0']:
                                        values.append(True)
                                    else:
                                        values.append(False)
                                # Преобразуем integer из float
                                elif not table_exists and 'INT' in column_types.get(col_name, ''):
                                    values.append(int(float(value)))
                                else:
                                    values.append(value)

                        if columns:
                            placeholders = ', '.join(['%s'] * len(values))
                            sql = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"
                            cursor.execute(sql, values)
                            inserted_count += 1

                    messages.success(request, f'Успешно импортировано {inserted_count} записей в таблицу {table_name}')
                    return redirect('table_management')

            except Exception as e:
                messages.error(request, f'Ошибка при восстановлении: {str(e)}')

    return render(request, 'core/table_restore_excel.html')


import numpy as np
from io import BytesIO
import re


def restore_database(request):
    """Умное восстановление БД из файлов"""
    if request.method == 'POST':
        restore_type = request.POST.get('restore_type')

        if restore_type == 'excel':
            # Восстановление из Excel
            excel_file = request.FILES.get('excel_file')
            restore_mode = request.POST.get('restore_mode', 'smart')

            if excel_file:
                try:
                    # Читаем Excel файл
                    xls = pd.ExcelFile(excel_file)
                    restored_tables = []
                    restored_records = {}
                    errors = []

                    with connection.cursor() as cursor:
                        # Если режим замены - сначала удаляем таблицы
                        if restore_mode == 'replace':
                            tables_to_drop = []
                            for sheet_name in xls.sheet_names:
                                if sheet_name != 'INFO':
                                    tables_to_drop.append(sheet_name)

                            for table in reversed(tables_to_drop):
                                try:
                                    cursor.execute(f"DROP TABLE IF EXISTS {table} CASCADE")
                                except:
                                    pass

                        for sheet_name in xls.sheet_names:
                            if sheet_name == 'INFO':
                                continue

                            try:
                                # Проверяем существование таблицы
                                cursor.execute("""
                                    SELECT EXISTS (
                                        SELECT 1 FROM information_schema.tables 
                                        WHERE table_schema = 'public' 
                                        AND table_name = %s
                                    )
                                """, [sheet_name])

                                table_exists = cursor.fetchone()[0]
                                df = pd.read_excel(xls, sheet_name)

                                # Конвертируем numpy типы в Python типы
                                def convert_value(value, col_name=None):
                                    """Конвертирует numpy и pandas типы в Python типы"""
                                    if pd.isna(value):
                                        return None
                                    elif isinstance(value, (np.int64, np.int32, np.int16)):
                                        return int(value)
                                    elif isinstance(value, (np.float64, np.float32)):
                                        if value == int(value):
                                            return int(value)
                                        return float(value)
                                    elif isinstance(value, np.bool_):
                                        return bool(value)
                                    elif isinstance(value, pd.Timestamp):
                                        return value.to_pydatetime()
                                    elif isinstance(value, str):
                                        # Проверяем, не является ли это неправильно сохраненным массивом
                                        # Если каждый символ через запятую - это баг
                                        if value.count(',') == len(value) // 2 and len(value) > 10:
                                            # Убираем запятые между буквами
                                            fixed_value = value.replace(', ', '')
                                            # Проверяем, не нужно ли это сохранить как массив
                                            if col_name and (
                                                    'crew' in col_name.lower() or 'members' in col_name.lower()):
                                                # Разбиваем по реальным разделителям (если они есть)
                                                if ';' in fixed_value:
                                                    items = fixed_value.split(';')
                                                elif ',' in fixed_value and not value.count(',') == len(value) // 2:
                                                    items = fixed_value.split(',')
                                                else:
                                                    # Считаем что это один элемент
                                                    items = [fixed_value]
                                                return '{' + ','.join(f'"{item.strip()}"' for item in items) + '}'
                                            return fixed_value
                                        # Проверяем на нормальный массив PostgreSQL
                                        elif value.startswith('{') and value.endswith('}'):
                                            return value
                                        # Проверяем, не список ли это через запятую
                                        elif col_name and ('crew' in col_name.lower() or 'members' in col_name.lower()):
                                            if ',' in value and not value.startswith('{'):
                                                items = value.split(',')
                                                return '{' + ','.join(f'"{item.strip()}"' for item in items) + '}'
                                        return value
                                    else:
                                        return value

                                if not table_exists:
                                    # Создаем таблицу с правильными ограничениями
                                    columns_sql = []
                                    column_types = {}
                                    primary_key = None

                                    for col in df.columns:
                                        col_name = col.lower()

                                        # Определяем тип данных
                                        sample = df[col].dropna()
                                        if sample.empty:
                                            col_type = "TEXT"
                                        else:
                                            first_val = sample.iloc[0]

                                            # Проверяем на массив или crew_members
                                            if col_name in ['crew_members', 'crew'] or 'crew' in col_name:
                                                col_type = "TEXT[]"
                                            elif isinstance(first_val, str) and first_val.startswith(
                                                    '{') and first_val.endswith('}'):
                                                col_type = "TEXT[]"
                                            elif sample.dtype == 'bool':
                                                col_type = "BOOLEAN"
                                            elif sample.dtype in ['int64', 'int32']:
                                                col_type = "INTEGER"
                                            elif sample.dtype in ['float64', 'float32']:
                                                if all(sample == sample.astype(int)):
                                                    col_type = "INTEGER"
                                                else:
                                                    col_type = "DECIMAL(10,2)"
                                            else:
                                                try:
                                                    pd.to_datetime(sample)
                                                    if any('T' in str(v) or ':' in str(v) for v in sample.head(10)):
                                                        col_type = "TIMESTAMP"
                                                    else:
                                                        col_type = "DATE"
                                                except:
                                                    max_len = sample.astype(str).str.len().max()
                                                    if max_len <= 255:
                                                        col_type = "VARCHAR(255)"
                                                    else:
                                                        col_type = "TEXT"

                                        column_types[col_name] = col_type

                                        # Определяем primary key
                                        if col_name == f"{sheet_name}_id" or (col_name == 'id' and not primary_key):
                                            columns_sql.append(f"{col_name} {col_type} PRIMARY KEY")
                                            primary_key = col_name
                                        else:
                                            columns_sql.append(f"{col_name} {col_type}")

                                    # Создаем таблицу
                                    create_sql = f"CREATE TABLE {sheet_name} ({', '.join(columns_sql)})"
                                    cursor.execute(create_sql)

                                    # Восстанавливаем foreign keys если возможно
                                    for col in df.columns:
                                        col_name = col.lower()
                                        if col_name.endswith('_id') and col_name != f"{sheet_name}_id":
                                            # Определяем таблицу по имени колонки
                                            ref_table = col_name[:-3]  # убираем _id

                                            # Проверяем существует ли таблица
                                            cursor.execute("""
                                                SELECT EXISTS (
                                                    SELECT 1 FROM information_schema.tables 
                                                    WHERE table_schema = 'public' 
                                                    AND table_name = %s
                                                )
                                            """, [ref_table])

                                            if cursor.fetchone()[0]:
                                                try:
                                                    alter_sql = f"""
                                                        ALTER TABLE {sheet_name} 
                                                        ADD CONSTRAINT fk_{sheet_name}_{col_name}
                                                        FOREIGN KEY ({col_name}) 
                                                        REFERENCES {ref_table}({col_name})
                                                        ON DELETE CASCADE
                                                    """
                                                    cursor.execute(alter_sql)
                                                except:
                                                    pass  # Игнорируем если не удалось создать FK

                                    restored_tables.append(sheet_name)

                                # Восстанавливаем данные
                                if restore_mode == 'replace' and table_exists:
                                    cursor.execute(f"TRUNCATE TABLE {sheet_name} CASCADE")

                                # Получаем primary key колонку
                                cursor.execute("""
                                    SELECT kcu.column_name
                                    FROM information_schema.table_constraints tc
                                    JOIN information_schema.key_column_usage kcu
                                        ON tc.constraint_name = kcu.constraint_name
                                    WHERE tc.table_name = %s 
                                    AND tc.constraint_type = 'PRIMARY KEY'
                                """, [sheet_name])

                                pk_result = cursor.fetchone()
                                pk_column = pk_result[0] if pk_result else None

                                # В умном режиме получаем ВСЕ существующие записи для проверки дубликатов
                                existing_records = set()
                                if restore_mode == 'smart' and table_exists:
                                    # Получаем все колонки для полной проверки
                                    cursor.execute(f"SELECT * FROM {sheet_name}")
                                    existing_data = cursor.fetchall()

                                    # Создаем хеш каждой записи для проверки полных дубликатов
                                    for row in existing_data:
                                        # Создаем строку из всех значений для сравнения
                                        row_hash = tuple(str(v) for v in row)
                                        existing_records.add(row_hash)

                                # Вставляем данные
                                inserted = 0
                                skipped = 0

                                for _, row in df.iterrows():
                                    # В умном режиме проверяем на полный дубликат
                                    if restore_mode == 'smart' and table_exists:
                                        # Создаем хеш текущей строки
                                        row_values = []
                                        for col in df.columns:
                                            value = convert_value(row[col], col.lower())
                                            row_values.append(str(value) if value is not None else 'None')

                                        row_hash = tuple(row_values)

                                        # Проверяем, не существует ли уже такая запись
                                        if row_hash in existing_records:
                                            skipped += 1
                                            continue

                                        # Также проверяем по primary key
                                        if pk_column:
                                            pk_col_upper = pk_column.upper()
                                            pk_col_lower = pk_column.lower()

                                            # Ищем значение PK в разных вариантах написания
                                            pk_value = None
                                            if pk_col_upper in row.index:
                                                pk_value = convert_value(row[pk_col_upper])
                                            elif pk_col_lower in row.index:
                                                pk_value = convert_value(row[pk_col_lower])
                                            elif pk_column in row.index:
                                                pk_value = convert_value(row[pk_column])

                                            if pk_value:
                                                # Проверяем существование по PK
                                                cursor.execute(f"SELECT 1 FROM {sheet_name} WHERE {pk_column} = %s",
                                                               [pk_value])
                                                if cursor.fetchone():
                                                    skipped += 1
                                                    continue

                                    columns = []
                                    values = []

                                    for col in df.columns:
                                        col_name = col.lower()
                                        value = row[col]

                                        if pd.notna(value):
                                            columns.append(col_name)
                                            converted_value = convert_value(value, col_name)
                                            values.append(converted_value)

                                    if columns:
                                        placeholders = ', '.join(['%s'] * len(values))
                                        sql = f"INSERT INTO {sheet_name} ({', '.join(columns)}) VALUES ({placeholders})"

                                        try:
                                            cursor.execute(sql, values)
                                            inserted += 1

                                            # Добавляем новую запись в existing_records
                                            if restore_mode == 'smart':
                                                row_values = []
                                                for col in df.columns:
                                                    value = convert_value(row[col], col.lower())
                                                    row_values.append(str(value) if value is not None else 'None')
                                                existing_records.add(tuple(row_values))

                                        except Exception as e:
                                            if 'duplicate key' in str(e):
                                                skipped += 1
                                            elif restore_mode != 'smart':
                                                raise e
                                            else:
                                                skipped += 1

                                restored_records[sheet_name] = {'inserted': inserted, 'skipped': skipped}

                            except Exception as e:
                                errors.append(f"{sheet_name}: {str(e)}")

                    # Формируем отчет
                    if restored_tables:
                        messages.success(request, f'Созданы таблицы: {", ".join(restored_tables)}')

                    for table, stats in restored_records.items():
                        if stats['inserted'] > 0:
                            msg = f'{table}: добавлено {stats["inserted"]} записей'
                            if stats['skipped'] > 0:
                                msg += f', пропущено {stats["skipped"]} (уже существуют)'
                            messages.info(request, msg)
                        elif stats['skipped'] > 0:
                            messages.info(request, f'{table}: все {stats["skipped"]} записей уже существуют')

                    if errors:
                        for error in errors:
                            messages.warning(request, f'Предупреждение: {error}')

                    if not errors or restored_records:
                        messages.success(request, 'Восстановление завершено!')

                except Exception as e:
                    messages.error(request, f'Критическая ошибка: {str(e)}')

        elif restore_type == 'sql':
            # SQL восстановление остается таким же, но добавим восстановление FK
            sql_file = request.FILES.get('sql_file')
            restore_mode = request.POST.get('restore_mode', 'smart')

            if sql_file:
                try:
                    sql_content = sql_file.read().decode('utf-8')

                    with connection.cursor() as cursor:
                        if restore_mode == 'replace':
                            cursor.execute("""
                                SELECT table_name 
                                FROM information_schema.tables 
                                WHERE table_schema = 'public' 
                                AND table_name NOT LIKE 'auth_%'
                                AND table_name NOT LIKE 'django_%'
                            """)

                            tables = [row[0] for row in cursor.fetchall()]
                            for table in reversed(tables):
                                try:
                                    cursor.execute(f"DROP TABLE IF EXISTS {table} CASCADE")
                                except:
                                    pass

                        # Выполняем SQL команды
                        commands = []
                        current_command = []

                        for line in sql_content.split('\n'):
                            if line.strip().startswith('--'):
                                continue

                            current_command.append(line)

                            if line.rstrip().endswith(';'):
                                commands.append('\n'.join(current_command))
                                current_command = []

                        if current_command:
                            commands.append('\n'.join(current_command))

                        success_count = 0
                        error_count = 0

                        for command in commands:
                            command = command.strip()
                            if not command or command.startswith('--'):
                                continue

                            try:
                                cursor.execute(command)
                                success_count += 1
                            except Exception as e:
                                error_count += 1
                                if restore_mode != 'smart':
                                    raise e

                        if error_count > 0 and restore_mode == 'smart':
                            messages.warning(request, f'Выполнено {success_count} команд, пропущено {error_count}')
                        else:
                            messages.success(request, f'База данных восстановлена! Выполнено {success_count} команд')

                except Exception as e:
                    messages.error(request, f'Ошибка при восстановлении из SQL: {str(e)}')

    return render(request, 'core/restore_database.html')

def universal_crud(request):
    """Выбор таблицы для универсальных операций"""
    tables = []

    with connection.cursor() as cursor:
        # Получаем все пользовательские таблицы
        cursor.execute("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = 'public' 
            AND table_name NOT LIKE 'auth_%'
            AND table_name NOT LIKE 'django_%'
            ORDER BY table_name
        """)

        for row in cursor.fetchall():
            table_name = row[0]
            # Получаем количество записей
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
                record_count = cursor.fetchone()[0]
            except:
                record_count = 0

            tables.append({
                'name': table_name,
                'records': record_count
            })

    return render(request, 'core/universal_crud.html', {'tables': tables})


def universal_crud_operation(request):
    """Универсальные CRUD операции для любой таблицы"""
    table_name = request.GET.get('table') or request.POST.get('table')
    operation = request.GET.get('operation') or request.POST.get('operation')

    if not table_name or not operation:
        return redirect('universal_crud')

    # Защита от SQL инъекций - проверяем существование таблицы
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name = %s
            )
        """, [table_name])

        if not cursor.fetchone()[0]:
            messages.error(request, 'Таблица не найдена')
            return redirect('universal_crud')

    # Обработка экспорта в Excel
    if operation == 'export':
        try:
            with connection.cursor() as cursor:
                # Используем безопасный запрос для экспорта
                cursor.execute(f"""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name = %s 
                    ORDER BY ordinal_position
                """, [table_name])

                columns_info = cursor.fetchall()

                # Строим запрос с обработкой дат
                select_parts = []
                for col_name, col_type in columns_info:
                    if col_type == 'date' or 'timestamp' in col_type:
                        # Обрабатываем даты как текст для избежания ошибок
                        select_parts.append(f"TO_CHAR({col_name}, 'DD.MM.YYYY') as {col_name}")
                    else:
                        select_parts.append(col_name)

                select_query = f"SELECT {', '.join(select_parts)} FROM {table_name}"
                cursor.execute(select_query)

                columns = [desc[0] for desc in cursor.description]
                data = cursor.fetchall()

                # Создаем DataFrame
                df = pd.DataFrame(data, columns=columns)

                # Создаем Excel файл в памяти
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=table_name[:31], index=False)

                    # Настройка ширины колонок
                    worksheet = writer.sheets[table_name[:31]]
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).map(len).max() if not df.empty else 0,
                            len(str(col))
                        ) + 2
                        worksheet.column_dimensions[
                            openpyxl.utils.get_column_letter(idx + 1)
                        ].width = min(max_length, 50)

                output.seek(0)

                # Отправляем файл
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                response['Content-Disposition'] = f'attachment; filename="{table_name}_{timestamp}.xlsx"'
                return response

        except Exception as e:
            messages.error(request, f'Ошибка при экспорте: {str(e)}')
            return redirect('universal_crud')

    context = {
        'table_name': table_name,
        'operation': operation
    }

    with connection.cursor() as cursor:
        # Получаем информацию о колонках с типами
        cursor.execute("""
            SELECT 
                column_name,
                data_type,
                is_nullable,
                column_default,
                udt_name
            FROM information_schema.columns
            WHERE table_name = %s
            ORDER BY ordinal_position
        """, [table_name])

        columns_info = cursor.fetchall()
        columns = []
        for col in columns_info:
            columns.append({
                'name': col[0],
                'type': col[1],
                'nullable': col[2] == 'YES',
                'default': col[3],
                'udt_name': col[4]
            })
        context['columns'] = columns

        # Обработка POST запросов
        if request.method == 'POST':
            action = request.POST.get('action')

            if action == 'add':
                # Добавление записи
                fields = []
                values = []
                placeholders = []

                # Находим primary key колонку
                cursor.execute("""
                    SELECT kcu.column_name
                    FROM information_schema.table_constraints tc
                    JOIN information_schema.key_column_usage kcu
                        ON tc.constraint_name = kcu.constraint_name
                    WHERE tc.table_name = %s 
                    AND tc.constraint_type = 'PRIMARY KEY'
                """, [table_name])

                pk_result = cursor.fetchone()
                pk_column = pk_result[0] if pk_result else None

                # Проверка на существующий ID, если он передан вручную
                if pk_column and pk_column in request.POST and request.POST[pk_column]:
                    check_id = request.POST[pk_column]
                    cursor.execute(f"SELECT EXISTS(SELECT 1 FROM {table_name} WHERE {pk_column} = %s)", [check_id])
                    if cursor.fetchone()[0]:
                        messages.error(request, f'Запись с ID {check_id} уже существует')
                        return redirect(f"{request.path}?table={table_name}&operation=add")

                    # Добавляем ID если он не существует
                    fields.append(pk_column)
                    values.append(check_id)
                    placeholders.append('%s')
                elif pk_column:
                    # Автогенерация ID
                    for col in columns:
                        if col['name'] == pk_column and ('int' in col['type'] or 'serial' in col['type']):
                            cursor.execute(f"SELECT COALESCE(MAX({pk_column}), 0) FROM {table_name}")
                            max_id = cursor.fetchone()[0]
                            new_id = max_id + 1
                            fields.append(pk_column)
                            values.append(new_id)
                            placeholders.append('%s')
                            break

                # Обрабатываем остальные поля
                for col in columns:
                    col_name = col['name']
                    col_type = col['type']

                    # Пропускаем primary key, если уже обработали
                    if col_name == pk_column and pk_column in fields:
                        continue

                    if col_name in request.POST:
                        value = request.POST[col_name]

                        if value:  # Если значение не пустое
                            # Обработка дат - преобразуем в формат YYYY-MM-DD
                            if col_type == 'date':
                                # input type="date" уже отправляет в формате YYYY-MM-DD
                                # но на всякий случай проверим
                                if '-' not in value and '.' in value:
                                    # Если пришло в формате DD.MM.YYYY, преобразуем
                                    try:
                                        parts = value.split('.')
                                        if len(parts) == 3:
                                            value = f"{parts[2]}-{parts[1]}-{parts[0]}"
                                    except:
                                        messages.error(request, f'Некорректный формат даты для поля {col_name}')
                                        return redirect(f"{request.path}?table={table_name}&operation=add")

                            # Обработка массивов
                            elif 'ARRAY' in col_type.upper():
                                # Преобразуем строку через запятую в PostgreSQL массив
                                array_values = [v.strip() for v in value.split(',') if v.strip()]
                                value = '{' + ','.join(
                                    f'"{v}"' if ' ' in v or '"' in v else v for v in array_values) + '}'

                            fields.append(col_name)
                            values.append(value)
                            placeholders.append('%s')
                        elif col['nullable']:
                            # Если поле может быть NULL
                            fields.append(col_name)
                            values.append(None)
                            placeholders.append('%s')

                if fields:
                    sql = f"INSERT INTO {table_name} ({', '.join(fields)}) VALUES ({', '.join(placeholders)})"
                    try:
                        cursor.execute(sql, values)
                        messages.success(request, 'Запись успешно добавлена')
                        return redirect(f"{request.path}?table={table_name}&operation=view")
                    except Exception as e:
                        messages.error(request, f'Ошибка при добавлении: {str(e)}')

            elif action == 'update':
                # Обновление записи
                record_id = request.POST.get('record_id')
                if record_id:
                    set_clause = []
                    values = []

                    for col in columns:
                        col_name = col['name']
                        col_type = col['type']

                        if col_name in request.POST and col_name != columns[0]['name']:  # Не обновляем ID
                            value = request.POST[col_name]

                            if value:
                                # Обработка дат
                                if col_type == 'date':
                                    if '-' not in value and '.' in value:
                                        try:
                                            parts = value.split('.')
                                            if len(parts) == 3:
                                                value = f"{parts[2]}-{parts[1]}-{parts[0]}"
                                        except:
                                            messages.error(request, f'Некорректный формат даты для поля {col_name}')
                                            return redirect(
                                                f"{request.path}?table={table_name}&operation=update&id={record_id}")

                                # Обработка массивов
                                elif 'ARRAY' in col_type.upper():
                                    array_values = [v.strip() for v in value.split(',') if v.strip()]
                                    value = '{' + ','.join(
                                        f'"{v}"' if ' ' in v or '"' in v else v for v in array_values) + '}'
                            else:
                                value = None

                            set_clause.append(f"{col_name} = %s")
                            values.append(value)

                    if set_clause:
                        values.append(record_id)
                        sql = f"UPDATE {table_name} SET {', '.join(set_clause)} WHERE {columns[0]['name']} = %s"
                        try:
                            cursor.execute(sql, values)
                            messages.success(request, 'Запись успешно обновлена')
                            return redirect(f"{request.path}?table={table_name}&operation=view")
                        except Exception as e:
                            messages.error(request, f'Ошибка при обновлении: {str(e)}')

            elif action == 'delete':
                # Удаление одной записи
                record_id = request.POST.get('record_id')
                if record_id:
                    sql = f"DELETE FROM {table_name} WHERE {columns[0]['name']} = %s"
                    try:
                        cursor.execute(sql, [record_id])
                        messages.success(request, 'Запись успешно удалена')
                    except Exception as e:
                        messages.error(request, f'Ошибка при удалении: {str(e)}')

            elif action == 'delete_multiple':
                # Удаление нескольких записей
                delete_ids = request.POST.getlist('delete_ids')
                if delete_ids:
                    placeholders = ', '.join(['%s'] * len(delete_ids))
                    sql = f"DELETE FROM {table_name} WHERE {columns[0]['name']} IN ({placeholders})"
                    try:
                        cursor.execute(sql, delete_ids)
                        messages.success(request, f'Удалено записей: {len(delete_ids)}')
                    except Exception as e:
                        messages.error(request, f'Ошибка при удалении: {str(e)}')

        # Обработка операций просмотра
        if operation == 'view' or operation == 'delete':
            try:
                # Фильтрация
                filter_column = request.GET.get('filter_column')
                filter_value = request.GET.get('filter_value')

                # Строим безопасный запрос с обработкой дат
                select_parts = []
                date_columns = []

                for i, col in enumerate(columns):
                    if col['type'] == 'date' or 'timestamp' in col['type']:
                        # Для дат используем TO_CHAR чтобы избежать ошибок
                        select_parts.append(f"TO_CHAR({col['name']}, 'YYYY-MM-DD') as {col['name']}")
                        date_columns.append(i)
                    else:
                        select_parts.append(col['name'])

                sql = f"SELECT {', '.join(select_parts)} FROM {table_name}"
                params = []

                if filter_column and filter_value and any(col['name'] == filter_column for col in columns):
                    sql += f" WHERE CAST({filter_column} AS TEXT) ILIKE %s"
                    params.append(f'%{filter_value}%')

                sql += f" ORDER BY {columns[0]['name']}"

                cursor.execute(sql, params)
                raw_data = cursor.fetchall()

                # Форматируем данные для шаблона
                data = []
                for row in raw_data:
                    formatted_row = {
                        'id': row[0],
                        'values': []
                    }
                    for i, value in enumerate(row):
                        col_type = columns[i]['type']
                        formatted_row['values'].append({
                            'value': value,
                            'type': col_type
                        })
                    data.append(formatted_row)

                context['data'] = data
                context['filter_column'] = filter_column
                context['filter_value'] = filter_value

            except Exception as e:
                messages.error(request, f'Ошибка при загрузке данных: {str(e)}')
                context['data'] = []

        elif operation == 'add':
            # Подготовка полей для добавления
            fields = []

            # Находим primary key
            cursor.execute("""
                SELECT kcu.column_name
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu
                    ON tc.constraint_name = kcu.constraint_name
                WHERE tc.table_name = %s 
                AND tc.constraint_type = 'PRIMARY KEY'
            """, [table_name])

            pk_result = cursor.fetchone()
            pk_column = pk_result[0] if pk_result else None

            for col in columns:
                # Для primary key integer показываем поле, но с подсказкой
                if pk_column and col['name'] == pk_column and ('int' in col['type'] or 'serial' in col['type']):
                    fields.append({
                        'name': col['name'],
                        'type': col['type'],
                        'nullable': False,
                        'is_pk': True,
                        'auto_increment': True
                    })
                else:
                    fields.append({
                        'name': col['name'],
                        'type': col['udt_name'] if 'ARRAY' in col['type'].upper() else col['type'],
                        'nullable': col['nullable'],
                        'is_pk': False,
                        'auto_increment': False
                    })
            context['fields'] = fields

        elif operation == 'update':
            # Получение записи для обновления
            record_id = request.GET.get('id')
            if record_id:
                try:
                    # Безопасный запрос с обработкой дат
                    select_parts = []
                    for col in columns:
                        if col['type'] == 'date' or 'timestamp' in col['type']:
                            select_parts.append(f"TO_CHAR({col['name']}, 'YYYY-MM-DD') as {col['name']}")
                        else:
                            select_parts.append(col['name'])

                    sql = f"SELECT {', '.join(select_parts)} FROM {table_name} WHERE {columns[0]['name']} = %s"
                    cursor.execute(sql, [record_id])
                    record = cursor.fetchone()

                    if record:
                        fields = []
                        for i, col in enumerate(columns):
                            # Не позволяем редактировать ID
                            if i == 0:
                                continue

                            fields.append({
                                'name': col['name'],
                                'type': col['udt_name'] if 'ARRAY' in col['type'].upper() else col['type'],
                                'nullable': col['nullable'],
                                'value': record[i] if record else None
                            })
                        context['fields'] = fields
                        context['record'] = record
                        context['record_id'] = record_id
                except Exception as e:
                    messages.error(request, f'Ошибка при загрузке записи: {str(e)}')

    return render(request, 'core/universal_crud_operation.html', context)